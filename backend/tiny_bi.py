from __future__ import annotations

import calendar
import json
import logging
import os
import sqlite3
import time
import unicodedata
from collections import Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from threading import Lock
from typing import Any, Optional

from openpyxl import load_workbook
import requests
from dotenv import load_dotenv

logger = logging.getLogger(__name__)

TINY_DIR = Path(__file__).resolve().parent
ROOT_DIR = TINY_DIR.parent

env_local = TINY_DIR / ".env"
env_parent = ROOT_DIR / ".env"
if env_local.exists():
    load_dotenv(env_local)
elif env_parent.exists():
    load_dotenv(env_parent)
else:
    load_dotenv()

DB_PATH = Path(os.getenv("TINY_DB_PATH") or (TINY_DIR / "db" / "tiny_bi.sqlite"))
SCHEMA_PATH = TINY_DIR / "db" / "schema_tiny_bi_sqlite.sql"
TINY_API_TOKEN = os.getenv("TINY_API_TOKEN")

TINY_BASE_URL = "https://api.tiny.com.br/api2"
REQUEST_TIMEOUT = int(os.getenv("TINY_REQUEST_TIMEOUT", "30"))
DETAIL_MAX_WORKERS = int(os.getenv("TINY_DETAIL_WORKERS", "6"))
DETAIL_RETRY_ATTEMPTS = int(os.getenv("TINY_DETAIL_RETRIES", "2"))
SYNC_LOOKBACK_DAYS = int(os.getenv("TINY_SYNC_LOOKBACK_DAYS", "7"))
INCREMENTAL_AUDIT_DAYS = int(os.getenv("TINY_INCREMENTAL_AUDIT_DAYS", "0"))
OPERATIONAL_WATCH_MAX_ORDERS = int(os.getenv("TINY_OPERATIONAL_WATCH_MAX_ORDERS", "200"))
TINY_MIN_REQUEST_INTERVAL = float(os.getenv("TINY_MIN_REQUEST_INTERVAL", "0.35"))
TINY_API_RATE_LIMIT_RETRIES = int(os.getenv("TINY_API_RATE_LIMIT_RETRIES", "3"))
TINY_API_RATE_LIMIT_SLEEP = float(os.getenv("TINY_API_RATE_LIMIT_SLEEP", "20"))
SQLITE_CONNECT_TIMEOUT = float(os.getenv("TINY_SQLITE_TIMEOUT_SECONDS", "30"))
SQLITE_BUSY_TIMEOUT_MS = int(os.getenv("TINY_SQLITE_BUSY_TIMEOUT_MS", "30000"))
SQLITE_CACHE_SIZE_KB = int(os.getenv("TINY_SQLITE_CACHE_SIZE_KB", "16384"))
SQLITE_MMAP_SIZE_MB = int(os.getenv("TINY_SQLITE_MMAP_SIZE_MB", "64"))
OPEN_ORDER_ALERT_DAYS = int(os.getenv("TINY_OPEN_ORDER_ALERT_DAYS", "1"))

MONTHS = ["JAN", "FEV", "MAR", "ABR", "MAI", "JUN", "JUL", "AGO", "SET", "OUT", "NOV", "DEZ"]
MONTH_HEADER_TOKEN_TO_INDEX = {
    "jan": 1,
    "janeiro": 1,
    "fev": 2,
    "fevereiro": 2,
    "mar": 3,
    "marco": 3,
    "abr": 4,
    "abril": 4,
    "mai": 5,
    "maio": 5,
    "jun": 6,
    "junho": 6,
    "jul": 7,
    "julho": 7,
    "ago": 8,
    "agosto": 8,
    "set": 9,
    "setembro": 9,
    "out": 10,
    "outubro": 10,
    "nov": 11,
    "novembro": 11,
    "dez": 12,
    "dezembro": 12,
}

STORE_RULES: dict[str, dict[str, Any]] = {
    "brew": {
        "label": "Brewhouse (E-commerce)",
        "type": "ecommerce",
        "ecommerce_name": "Brewhouse",
    },
    "grow": {
        "label": "Growhouse (E-commerce)",
        "type": "ecommerce",
        "ecommerce_name": "Growhouse",
    },
    "brewnh": {
        "label": "Brew NH (Loja Física)",
        "type": "tag",
        "tags": {"brewnh"},
        "marker_tag": "brewnh",
    },
    "brewpoa": {
        "label": "Brew POA (Loja Física)",
        "type": "tag",
        "tags": {"brewpoa"},
        "marker_tag": "brewpoa",
    },
    "bigb": {
        "label": "Big B",
        "type": "tag",
        "tags": {"bigb"},
        "marker_tag": "bigb",
    },
    "grow_fisica": {
        "label": "Growhouse (Loja Física)",
        "type": "tag",
        "tags": {"grow"},
        "marker_tag": "grow",
    },
}

DEFAULT_ECOM_TARGET_FILES: dict[str, Path] = {
    "brew": TINY_DIR / "planilhas" / "brew" / "Metas_Realizado_2021_2026.xlsx",
    "grow": TINY_DIR / "planilhas" / "grow" / "Metas_Realizado_2025_2026.xlsx",
}
DEFAULT_BREW_HISTORY_FILE = TINY_DIR.parent / "Faturamento E-commerce Brew - 2021-2026.xlsx"

PRODUCT_PERIODS = {
    "month": "Mes atual",
    "prev_month": "Mes anterior",
    "current_year": "Ano atual",
    "previous_year": "Ano anterior",
    "30d": "Ultimos 30 dias",
    "6m": "Ultimos 6 meses",
    "12m": "Ultimos 12 meses",
}

PRODUCT_METRICS = {"qty", "revenue"}

CUSTOMER_PERIODS = {
    "month": "Mes atual",
    "prev_month": "Mes anterior",
    "current_year": "Ano atual",
    "previous_year": "Ano anterior",
    "30d": "Ultimos 30 dias",
    "6m": "Ultimos 6 meses",
    "12m": "Ultimos 12 meses",
    "custom": "Período personalizado",
}

FUNNEL_PERIODS = {
    "7d": "Ultimos 7 dias",
    "month": "Mes atual",
    "prev_month": "Mes anterior",
    "current_year": "Ano atual",
    "previous_year": "Ano anterior",
    "30d": "Ultimos 30 dias",
    "90d": "Ultimos 90 dias",
    "6m": "Ultimos 6 meses",
    "12m": "Ultimos 12 meses",
}

MANAGEMENT_PERIODS = {
    "7d": "Ultimos 7 dias",
    "month": "Mes atual",
    "prev_month": "Mes anterior",
    "current_year": "Ano atual",
    "previous_year": "Ano anterior",
    "30d": "Ultimos 30 dias",
    "90d": "Ultimos 90 dias",
    "6m": "Ultimos 6 meses",
    "12m": "Ultimos 12 meses",
}

TINY_SHIPPING_METHOD_LABELS = {
    "S": "Sem Frete",
    "X": "Customizada",
}

FUNNEL_STEP_DEFINITIONS = [
    {
        "key": "sessions",
        "label": "Visitas totais",
        "description": "Total de sessões que chegaram ao site no período.",
        "focus": "Topo de funil",
    },
    {
        "key": "engaged_sessions",
        "label": "Sessões engajadas",
        "description": "Sessões com interação ativa (tempo, páginas ou evento).",
        "focus": "Qualidade de tráfego",
    },
    {
        "key": "add_to_cart",
        "label": "Adições ao carrinho",
        "description": "Sessões em que houve inclusão de item no carrinho.",
        "focus": "Intenção de compra",
    },
    {
        "key": "purchase",
        "label": "Pedidos finalizados",
        "description": "Eventos de compra confirmada no período.",
        "focus": "Conversao final",
    },
]

# Maps GA4 sessionDefaultChannelGroup → (internal_key, label, color)
# "Cross-network" = Performance Max; "Display" = Google Display Network
CHANNEL_GROUPS: dict[str, tuple[str, str, str]] = {
    "Organic Search":   ("organico",   "Orgânico",     "#38bdf8"),
    "Paid Search":      ("google_ads", "Google Ads",   "#34d399"),
    "Cross-network":    ("google_ads", "Google Ads",   "#34d399"),
    "Display":          ("google_ads", "Google Ads",   "#34d399"),
    "Paid Social":      ("meta_ads",   "Meta Ads",     "#c084fc"),
    "Organic Social":   ("social_org", "Social (org)", "#fb923c"),
    "Email":            ("email",      "E-mail",       "#f472b6"),
}
CHANNEL_ORDER = ["organico", "google_ads", "meta_ads", "social_org", "email", "outros"]
CHANNEL_META: dict[str, tuple[str, str]] = {
    "organico":   ("Orgânico",     "#38bdf8"),
    "google_ads": ("Google Ads",   "#34d399"),
    "meta_ads":   ("Meta Ads",     "#c084fc"),
    "social_org": ("Social (org)", "#fb923c"),
    "email":      ("E-mail",       "#f472b6"),
    "outros":     ("Outros",       "#475569"),
}

GA4_PROPERTY_ID_BREW = os.getenv("GA4_PROPERTY_ID_BREW") or os.getenv("GA4_PROPERTY_ID")
GA4_PROPERTY_ID_GROW = os.getenv("GA4_PROPERTY_ID_GROW")
GA4_SCOPES = [
    scope.strip()
    for scope in str(
        os.getenv("GA4_SCOPES") or "https://www.googleapis.com/auth/analytics.readonly"
    ).split(",")
    if scope.strip()
]
GOOGLE_APPLICATION_CREDENTIALS = os.getenv("GOOGLE_APPLICATION_CREDENTIALS")

AGE_BUCKETS = [
    ("Até 17", 0, 17),
    ("18-24", 18, 24),
    ("25-34", 25, 34),
    ("35-44", 35, 44),
    ("45-54", 45, 54),
    ("55-64", 55, 64),
    ("65+", 65, None),
]

LOYALTY_BUCKETS = [
    ("1 pedido", 1, 1),
    ("2-3 pedidos", 2, 3),
    ("4-6 pedidos", 4, 6),
    ("7+ pedidos", 7, None),
]

EXCLUDED_ORDER_SITUACAO_TOKENS = tuple(
    token.strip().lower()
    for token in str(os.getenv("TINY_EXCLUDED_SITUACOES") or "cancelado,em aberto").split(",")
    if token.strip()
)
EXCLUDED_REVENUE_MARKER_VALUES = tuple(
    token.strip()
    for token in str(os.getenv("TINY_EXCLUDED_REVENUE_MARKERS") or "racao,ração").split(",")
    if token.strip()
)


_REQUEST_LOCK = Lock()
_LAST_REQUEST_AT = 0.0
_DB_INIT_LOCK = Lock()
_DB_READY = False
_STALE_SYNC_RUNS_REPAIRED = False


def _now_utc_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def _normalize_text(value: Optional[str]) -> str:
    return (value or "").strip().lower()


def _parse_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        parsed = float(value)
        if parsed != parsed:  # NaN
            return None
        return parsed
    if isinstance(value, str):
        normalized = value.strip()
        if not normalized:
            return None
        if "," in normalized and "." in normalized:
            normalized = normalized.replace(".", "").replace(",", ".")
        elif "," in normalized:
            normalized = normalized.replace(",", ".")
        try:
            parsed = float(normalized)
            if parsed != parsed:  # NaN
                return None
            return parsed
        except ValueError:
            return None
    return None


def _parse_int(value: Any) -> Optional[int]:
    if value is None:
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        if value != value:  # NaN
            return None
        return int(value)
    if isinstance(value, str):
        normalized = value.strip()
        if not normalized:
            return None
        normalized = normalized.replace(",", ".")
        try:
            return int(float(normalized))
        except ValueError:
            return None
    return None


def _parse_date(value: Any) -> Optional[date]:
    if not value:
        return None
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        normalized = value.strip()
        if not normalized:
            return None
        candidates = [normalized]
        if "." in normalized:
            candidates.append(normalized.split(".", 1)[0])
        if len(normalized) >= 19:
            candidates.append(normalized[:19])
        if len(normalized) >= 10:
            candidates.append(normalized[:10])

        for candidate in dict.fromkeys(candidates):
            for fmt in (
                "%d/%m/%Y",
                "%m/%d/%Y",
                "%m/%d/%Y %H:%M:%S",
                "%Y-%m-%d",
                "%Y-%m-%d %H:%M:%S",
                "%Y-%m-%dT%H:%M:%S",
            ):
                try:
                    return datetime.strptime(candidate, fmt).date()
                except ValueError:
                    continue
    return None


def _normalize_token(value: Any) -> str:
    normalized = unicodedata.normalize("NFKD", str(value or ""))
    ascii_value = normalized.encode("ascii", "ignore").decode("ascii").lower()
    return "".join(ch for ch in ascii_value if ch.isalnum())


EXCLUDED_REVENUE_MARKER_TOKENS = {
    _normalize_token(token)
    for token in EXCLUDED_REVENUE_MARKER_VALUES
    if _normalize_token(token)
}


KNOWN_MARKER_TAG_TOKENS = {
    _normalize_token(tag)
    for rule in STORE_RULES.values()
    if rule.get("type") == "tag"
    for tag in (rule.get("tags") or set())
    if _normalize_token(tag)
}


def _markers_from_text(markers_text: Optional[str]) -> list[str]:
    if not markers_text:
        return []
    return [chunk.strip() for chunk in str(markers_text).split(",") if chunk and chunk.strip()]


def _is_excluded_revenue_marker(marker: Any) -> bool:
    raw = str(marker or "").strip().lower()
    if not raw:
        return False
    token = _normalize_token(raw)
    if token and token in EXCLUDED_REVENUE_MARKER_TOKENS:
        return True
    # Alguns marcadores antigos vieram com caractere de substituicao (ex.: "ra��o").
    if "�" in raw and token == "rao" and "racao" in EXCLUDED_REVENUE_MARKER_TOKENS:
        return True
    return False


def _has_excluded_revenue_marker(markers: list[str]) -> bool:
    return any(_is_excluded_revenue_marker(marker) for marker in markers)


def _contains_em_aberto(situacao: Optional[str]) -> bool:
    token = _normalize_token(situacao)
    return "emaberto" in token if token else False


def _pickup_location_label(raw_order: dict[str, Any]) -> Optional[str]:
    for key in ("forma_frete", "formaFrete", "obs", "obs_interna", "obsInterna"):
        raw_value = str(raw_order.get(key) or "").strip().lower()
        if not raw_value:
            continue
        if "loja nh" in raw_value or "retirar na loja nh" in raw_value:
            return "Retirada em NH"
        if "loja poa" in raw_value or "retirar na loja poa" in raw_value:
            return "Retirada em POA"
    return None


def _normalize_shipping_label(label: Optional[str], raw_order: dict[str, Any]) -> Optional[str]:
    raw_label = str(label or "").strip()
    normalized = _normalize_token(raw_label)

    if not raw_label:
        return None
    if "saomiguel" in normalized:
        return "São Miguel"
    if any(token in normalized for token in ("texcourier", "tntmercurio", "totalexpress")):
        return "Total Express"
    if any(token in normalized for token in ("chrservspostais", "correios", "sedex", "pac")):
        return "Correios"
    return raw_label


def _normalize_payment_label(value: Any) -> Optional[str]:
    raw_label = str(value or "").strip()
    if not raw_label:
        return None

    normalized = _normalize_token(raw_label)
    if normalized == "pix":
        return "PIX"
    if normalized == "credito":
        return "Crédito"
    if normalized == "debito":
        return "Débito"
    if normalized == "boleto":
        return "Boleto"
    if normalized == "crediario":
        return "Crediário"
    if normalized.startswith("dinheiro"):
        return "Dinheiro"
    if normalized == "multiplas":
        return "Múltiplas"
    if normalized == "deposito":
        return "Depósito"
    if normalized == "cheque":
        return "Cheque"
    return raw_label


def _extract_shipping_method(raw_order: dict[str, Any]) -> tuple[Optional[str], Optional[str]]:
    raw_code = str(raw_order.get("forma_envio") or raw_order.get("formaEnvio") or "").strip().upper()
    shipping_code = raw_code or None
    transporter = str(
        raw_order.get("nome_transportador")
        or raw_order.get("nomeTransportador")
        or raw_order.get("transportador")
        or ""
    ).strip()

    if transporter:
        return shipping_code, _normalize_shipping_label(transporter, raw_order)

    pickup_label = _pickup_location_label(raw_order)
    if pickup_label:
        return shipping_code, pickup_label

    if shipping_code in {"S", "X"}:
        return shipping_code, "Não especificado"

    # O código X do Tiny é genérico demais para análise. Quando houver nome
    # de transportadora, ele deve prevalecer para melhorar filtros e leitura.
    label = TINY_SHIPPING_METHOD_LABELS.get(raw_code)
    normalized_label = _normalize_shipping_label(label, raw_order)
    if normalized_label:
        return shipping_code, normalized_label

    if label:
        return shipping_code, label

    if shipping_code:
        return shipping_code, _normalize_shipping_label(f"Código {shipping_code}", raw_order)
    return None, None


def _shipping_labels_need_refresh(conn: sqlite3.Connection) -> bool:
    try:
        row = conn.execute("SELECT COUNT(*) AS count FROM orders WHERE COALESCE(trim(raw_payload), '') <> ''").fetchone()
        return int((row["count"] if row else 0) or 0) > 0
    except sqlite3.OperationalError as exc:
        if "no such table" in str(exc).lower():
            return False
        raise


def _compute_operational_flags(
    *,
    order_date: Optional[date],
    created_at: Optional[date],
    situacao: Optional[str],
    numero_ecommerce: Any,
    ecommerce_name: Any,
    markers: list[str],
) -> dict[str, int]:
    has_ecommerce_link = int(bool(str(numero_ecommerce or "").strip() or str(ecommerce_name or "").strip()))
    marker_tokens = {_normalize_token(marker) for marker in markers if _normalize_token(marker)}
    has_excluded_revenue_marker = _has_excluded_revenue_marker(markers)
    has_known_marker = int(bool(marker_tokens.intersection(KNOWN_MARKER_TAG_TOKENS)) or has_excluded_revenue_marker)
    # Pedidos com marcador excluído de receita (ex: "racao") têm has_known_marker=1 por design —
    # são "identificáveis" operacionalmente, apenas excluídos da receita.
    # Por isso não entram no alerta de "sem loja identificada". Comportamento correto e intencional.
    flag_missing_link_and_marker = int((has_ecommerce_link == 0) and (has_known_marker == 0))

    base_date = created_at or order_date
    flag_open_over_n_days = 0
    if (not has_excluded_revenue_marker) and _contains_em_aberto(situacao) and isinstance(base_date, date):
        flag_open_over_n_days = int((date.today() - base_date).days > OPEN_ORDER_ALERT_DAYS)

    return {
        "hasEcommerceLink": has_ecommerce_link,
        "hasKnownMarker": has_known_marker,
        "hasExcludedRevenueMarker": int(has_excluded_revenue_marker),
        "flagMissingLinkAndMarker": flag_missing_link_and_marker,
        "flagOpenOverNDays": flag_open_over_n_days,
    }


def _month_from_header(value: Any) -> Optional[int]:
    token = _normalize_token(value)
    if token in MONTH_HEADER_TOKEN_TO_INDEX:
        return MONTH_HEADER_TOKEN_TO_INDEX[token]
    return None


def _target_row_type(value: Any) -> Optional[str]:
    token = _normalize_token(value)
    if not token:
        return None
    if "meta" in token:
        return "target"
    if "realizado" in token or token in {"real", "realizacao"}:
        return "realized"
    return None


def _week_segments(year: int, month: int) -> list[tuple[int, date, date]]:
    month_start = date(year, month, 1)
    month_end = date(year, month, calendar.monthrange(year, month)[1])
    week_start = month_start - timedelta(days=month_start.weekday())

    segments: list[tuple[int, date, date]] = []
    week_index = 1
    while week_start <= month_end:
        week_end = week_start + timedelta(days=6)
        segments.append((week_index, max(month_start, week_start), min(month_end, week_end)))
        week_index += 1
        week_start += timedelta(days=7)
    return segments


def _shift_months(anchor: datetime, months: int) -> datetime:
    month_index = anchor.month - 1 + months
    year = anchor.year + month_index // 12
    month = month_index % 12 + 1
    day = min(anchor.day, calendar.monthrange(year, month)[1])
    return anchor.replace(year=year, month=month, day=day)


def _period_range(period: str, now: datetime) -> tuple[datetime, datetime]:
    if period == "7d":
        return now - timedelta(days=7), now
    if period == "month":
        start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        return start, now
    if period == "prev_month":
        current_month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        prev_month_end = current_month_start - timedelta(seconds=1)
        prev_month_start = prev_month_end.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        return prev_month_start, prev_month_end
    if period == "current_year":
        start = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
        return start, now
    if period == "previous_year":
        start = now.replace(year=now.year - 1, month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
        end = start.replace(month=12, day=31, hour=23, minute=59, second=59, microsecond=0)
        return start, end
    if period == "30d":
        return now - timedelta(days=30), now
    if period == "90d":
        return now - timedelta(days=90), now
    if period == "6m":
        return _shift_months(now, -6), now
    if period == "12m":
        return _shift_months(now, -12), now
    raise ValueError("Periodo invalido")


def _parse_iso_date_input(value: str, field_name: str) -> date:
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError as exc:
        raise ValueError(f"{field_name} invalido. Use YYYY-MM-DD.") from exc


def _resolve_period_dates(
    period: str,
    period_labels: dict[str, str],
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    now: Optional[datetime] = None,
) -> tuple[date, date, str, str, bool]:
    if period not in period_labels:
        raise ValueError("Periodo invalido.")

    has_start = bool(start_date)
    has_end = bool(end_date)
    if has_start != has_end:
        raise ValueError("Para intervalo personalizado, informe start e end.")

    if has_start and has_end:
        start = _parse_iso_date_input(str(start_date), "start")
        end = _parse_iso_date_input(str(end_date), "end")
        if end < start:
            raise ValueError("end nao pode ser menor que start.")
        label = f"{start.isoformat()} a {end.isoformat()}"
        return start, end, "custom", label, True

    anchor = now or datetime.now(timezone.utc).replace(tzinfo=None)
    start_dt, end_dt = _period_range(period, anchor)
    return start_dt.date(), end_dt.date(), period, period_labels[period], False


def _distribution_from_counter(
    counter: Counter,
    *,
    top: Optional[int] = None,
    include_other: bool = False,
) -> list[dict[str, Any]]:
    if not counter:
        return []
    items = counter.most_common(top)
    if include_other and top is not None:
        others = sum(counter.values()) - sum(value for _, value in items)
        if others > 0:
            items.append(("Outros", others))
    return [{"label": label, "count": value} for label, value in items if value > 0]


def _orders_status_filter_sql(alias: Optional[str] = None) -> str:
    prefix = f"{alias}." if alias else ""
    column = f"lower(trim(coalesce({prefix}situacao, '')))"
    parts = [f"{column} NOT LIKE '%{token}%'" for token in EXCLUDED_ORDER_SITUACAO_TOKENS]
    return " AND ".join(parts)


def _orders_analytics_filter_sql(alias: Optional[str] = None) -> str:
    prefix = f"{alias}." if alias else ""
    status_filter = _orders_status_filter_sql(alias)
    revenue_filter = f"COALESCE({prefix}exclude_from_revenue, 0) = 0"
    return f"{status_filter} AND {revenue_filter}" if status_filter else revenue_filter


def _source_scope_sql(source: str, alias: Optional[str] = None) -> tuple[str, tuple[Any, ...]]:
    if source == "all":
        return "1 = 1", ()
    if source not in STORE_RULES:
        raise ValueError("Fonte desconhecida.")
    prefix = f"{alias}." if alias else ""
    return f"{prefix}store_code = ?", (source,)


def _bucket_value(value: Optional[int], buckets: list[tuple[str, int, Optional[int]]]) -> Optional[str]:
    if value is None:
        return None
    for label, start, end in buckets:
        if end is None and value >= start:
            return label
        if end is not None and start <= value <= end:
            return label
    return None


def _tiny_get(endpoint: str, params: dict[str, Any]) -> tuple[Optional[dict[str, Any]], Optional[str]]:
    if not TINY_API_TOKEN:
        return None, "TINY_API_TOKEN nao configurado."

    url = f"{TINY_BASE_URL}/{endpoint}"
    request_params = {"token": TINY_API_TOKEN, "formato": "JSON", **params}

    for attempt in range(TINY_API_RATE_LIMIT_RETRIES + 1):
        _throttle_tiny_request()
        try:
            response = requests.get(url, params=request_params, timeout=REQUEST_TIMEOUT)
            response.raise_for_status()
        except requests.RequestException as exc:
            if attempt < TINY_API_RATE_LIMIT_RETRIES:
                time.sleep(min(2.0 * (attempt + 1), 5.0))
                continue
            return None, f"Falha na API Tiny: {exc}"

        try:
            payload = response.json()
        except ValueError:
            return None, "Resposta invalida da API Tiny."

        retorno = payload.get("retorno")
        if not retorno:
            return None, "Resposta sem campo retorno."

        if retorno.get("status") == "OK":
            return retorno, None

        error = _tiny_error_message(retorno)
        if _is_rate_limited_error(error) and attempt < TINY_API_RATE_LIMIT_RETRIES:
            time.sleep(TINY_API_RATE_LIMIT_SLEEP * (attempt + 1))
            continue
        return None, error

    return None, "Erro na API Tiny."


def _tiny_post(endpoint: str, params: dict[str, Any]) -> tuple[Optional[dict[str, Any]], Optional[str]]:
    if not TINY_API_TOKEN:
        return None, "TINY_API_TOKEN nao configurado."

    url = f"{TINY_BASE_URL}/{endpoint}"
    request_params = {"token": TINY_API_TOKEN, "formato": "JSON", **params}

    for attempt in range(TINY_API_RATE_LIMIT_RETRIES + 1):
        _throttle_tiny_request()
        try:
            response = requests.post(url, data=request_params, timeout=REQUEST_TIMEOUT)
            response.raise_for_status()
        except requests.RequestException as exc:
            if attempt < TINY_API_RATE_LIMIT_RETRIES:
                time.sleep(min(2.0 * (attempt + 1), 5.0))
                continue
            return None, f"Falha na API Tiny: {exc}"

        try:
            payload = response.json()
        except ValueError:
            return None, "Resposta invalida da API Tiny."

        retorno = payload.get("retorno")
        if not retorno:
            return None, "Resposta sem campo retorno."

        if retorno.get("status") == "OK":
            return retorno, None

        error = _tiny_error_message(retorno)
        if _is_rate_limited_error(error) and attempt < TINY_API_RATE_LIMIT_RETRIES:
            time.sleep(TINY_API_RATE_LIMIT_SLEEP * (attempt + 1))
            continue
        return None, error

    return None, "Erro na API Tiny."


def _throttle_tiny_request() -> None:
    global _LAST_REQUEST_AT
    if TINY_MIN_REQUEST_INTERVAL <= 0:
        return
    with _REQUEST_LOCK:
        now = time.monotonic()
        wait_for = TINY_MIN_REQUEST_INTERVAL - (now - _LAST_REQUEST_AT)
        if wait_for > 0:
            time.sleep(wait_for)
        _LAST_REQUEST_AT = time.monotonic()


def _is_rate_limited_error(error: Optional[str]) -> bool:
    normalized = unicodedata.normalize("NFKD", (error or "").lower()).encode("ascii", "ignore").decode("ascii")
    return ("api bloqueada" in normalized) or ("excedido o numero de acessos" in normalized)


def _tiny_error_message(retorno: dict[str, Any]) -> str:
    if retorno.get("erro"):
        return str(retorno.get("erro"))

    errors = retorno.get("erros")
    messages: list[str] = []

    if isinstance(errors, list):
        for entry in errors:
            if isinstance(entry, dict):
                msg = entry.get("erro") or entry.get("mensagem") or entry.get("message")
                if msg:
                    messages.append(str(msg))
            elif entry:
                messages.append(str(entry))
    elif isinstance(errors, dict):
        msg = errors.get("erro") or errors.get("mensagem") or errors.get("message")
        if msg:
            messages.append(str(msg))
    elif errors:
        messages.append(str(errors))

    if messages:
        return " | ".join(messages)

    code = retorno.get("codigo_erro")
    if code is not None:
        return f"Erro na API Tiny (codigo {code})."
    return "Erro na API Tiny."


def _extract_markers(raw: Any) -> list[str]:
    markers: list[str] = []

    def add(value: Any) -> None:
        if value is None:
            return
        if isinstance(value, dict):
            for key in ("marcador", "descricao", "nome", "tag"):
                if key in value:
                    add(value[key])
                    return
            markers.append(str(value).strip().lower())
            return
        if isinstance(value, list):
            for item in value:
                add(item)
            return
        markers.append(str(value).strip().lower())

    add(raw)
    return [m for m in markers if m]


def _normalize_order(pedido: dict[str, Any]) -> dict[str, Any]:
    ecommerce = pedido.get("ecommerce") or {}
    ecommerce_name = None
    if isinstance(ecommerce, dict):
        ecommerce_name = ecommerce.get("nomeEcommerce") or ecommerce.get("nome")

    numero_ecommerce = pedido.get("numero_ecommerce")
    if not numero_ecommerce and isinstance(ecommerce, dict):
        numero_ecommerce = ecommerce.get("numeroPedidoEcommerce") or ecommerce.get("numero_pedido")

    situacao_payload = pedido.get("situacao") or {}
    if isinstance(situacao_payload, dict):
        situacao_id = situacao_payload.get("id") or situacao_payload.get("codigo")
        situacao = (
            situacao_payload.get("descricao")
            or situacao_payload.get("situacao")
            or situacao_payload.get("nome")
        )
    else:
        situacao_id = None
        situacao = situacao_payload

    customer = pedido.get("cliente") or {}
    if not isinstance(customer, dict):
        customer = {}

    payment = pedido.get("forma_pagamento") or pedido.get("formaPagamento") or {}
    payment_name = payment.get("nome") if isinstance(payment, dict) else payment
    payment_name = _normalize_payment_label(payment_name)
    shipping_method_code, shipping_method_label = _extract_shipping_method(pedido)

    items: list[dict[str, Any]] = []
    for idx, entry in enumerate(pedido.get("itens") or [], start=1):
        item = entry.get("item") if isinstance(entry, dict) else entry
        if not isinstance(item, dict):
            continue
        qty = _parse_float(item.get("quantidade")) or 0.0
        unit_price = _parse_float(item.get("valor_unitario"))
        line_total = _parse_float(item.get("valor_total"))
        revenue = line_total
        if revenue is None:
            revenue = (unit_price or 0.0) * qty

        items.append(
            {
                "line_number": idx,
                "sku": item.get("codigo") or item.get("sku"),
                "product_id": item.get("id_produto") or item.get("idProduto"),
                "name": item.get("descricao") or item.get("nome"),
                "quantity": qty,
                "unit_price": unit_price,
                "revenue": revenue,
                "raw": item,
            }
        )

    return {
        "tiny_id": pedido.get("id"),
        "number": pedido.get("numero"),
        "date": _parse_date(pedido.get("data_pedido") or pedido.get("dataPedido")),
        "created_at": _parse_date(pedido.get("data_criacao") or pedido.get("dataCriacao")),
        "updated_at": _parse_date(pedido.get("data_atualizacao") or pedido.get("dataAtualizacao")),
        "total": _parse_float(pedido.get("total_pedido") or pedido.get("valor_total")) or 0.0,
        "subtotal": _parse_float(pedido.get("valor_produtos") or pedido.get("subtotal")),
        "shipping": _parse_float(pedido.get("valor_frete") or pedido.get("frete")),
        "discount": _parse_float(pedido.get("valor_desconto") or pedido.get("desconto")),
        "situacao_id": int(situacao_id) if str(situacao_id).isdigit() else None,
        "situacao": str(situacao or "").strip() or None,
        "ecommerce_name": ecommerce_name,
        "numero_ecommerce": numero_ecommerce,
        "markers": _extract_markers(pedido.get("marcadores")),
        "payment_method": payment_name,
        "shipping_method_code": shipping_method_code,
        "shipping_method_label": shipping_method_label,
        "customer_code": customer.get("codigo"),
        "customer_name": customer.get("nome"),
        "customer_email": customer.get("email"),
        "customer_document": customer.get("cpf_cnpj") or customer.get("cpfCnpj"),
        "customer_city": customer.get("cidade"),
        "customer_state": customer.get("uf"),
        "items": items,
    }


def _store_match(order: dict[str, Any], source: str) -> bool:
    rule = STORE_RULES.get(source)
    if not rule:
        return False

    if rule.get("type") == "ecommerce":
        expected = _normalize_text(rule.get("ecommerce_name"))
        ecommerce_name = _normalize_text(order.get("ecommerce_name"))
        if ecommerce_name and expected:
            return ecommerce_name == expected
        return False

    if rule.get("type") == "tag":
        tags = {str(tag).lower() for tag in (rule.get("tags") or set())}
        markers = {str(marker).lower() for marker in (order.get("markers") or [])}
        return bool(tags.intersection(markers))

    return False


def _fetch_order_summaries(start: date, end: date, source: str) -> tuple[list[dict[str, Any]], Optional[str]]:
    summaries: list[dict[str, Any]] = []
    page = 1

    marker = None
    rule = STORE_RULES.get(source) or {}
    if rule.get("type") == "tag":
        marker = rule.get("marker_tag")

    while True:
        params: dict[str, Any] = {
            "pagina": page,
            "dataInicial": start.strftime("%d/%m/%Y"),
            "dataFinal": end.strftime("%d/%m/%Y"),
        }
        if marker:
            params["marcador"] = marker

        retorno, error = _tiny_post("pedidos.pesquisa.php", params)
        if error:
            normalized_error = unicodedata.normalize("NFKD", error.lower()).encode("ascii", "ignore").decode("ascii")
            if "consulta nao retornou registros" in normalized_error:
                return [], None
            return [], error

        pedidos = retorno.get("pedidos") or []
        for entry in pedidos:
            pedido = entry.get("pedido") if isinstance(entry, dict) else None
            if isinstance(pedido, dict):
                summaries.append(pedido)

        total_pages = int(retorno.get("numero_paginas") or 1)
        if page >= total_pages:
            break
        page += 1

    return summaries, None


def _fetch_order_detail(order_id: int) -> tuple[Optional[dict[str, Any]], Optional[str]]:
    retorno, error = _tiny_get("pedido.obter.php", {"id": order_id})
    if error:
        return None, error
    pedido = retorno.get("pedido")
    if not isinstance(pedido, dict):
        return None, "Detalhe do pedido invalido."
    return pedido, None


def _fetch_order_detail_with_retry(
    order_id: int,
    retries: int = DETAIL_RETRY_ATTEMPTS,
) -> tuple[Optional[dict[str, Any]], Optional[str]]:
    last_error = None
    for attempt in range(retries + 1):
        detail, error = _fetch_order_detail(order_id)
        if not error and detail:
            return detail, None
        last_error = error or "Falha ao obter detalhe do pedido."
        if attempt < retries:
            time.sleep(0.35 * (attempt + 1))
    return None, last_error


def _db_connect() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH, timeout=SQLITE_CONNECT_TIMEOUT)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    conn.execute("PRAGMA journal_mode = WAL")
    conn.execute("PRAGMA synchronous = NORMAL")
    conn.execute(f"PRAGMA busy_timeout = {max(1000, SQLITE_BUSY_TIMEOUT_MS)}")
    conn.execute("PRAGMA temp_store = MEMORY")
    conn.execute(f"PRAGMA cache_size = {-max(2048, SQLITE_CACHE_SIZE_KB)}")
    conn.execute(f"PRAGMA mmap_size = {max(0, SQLITE_MMAP_SIZE_MB) * 1024 * 1024}")
    return conn


def _safe_commit(
    conn: sqlite3.Connection,
    *,
    retries: int = 4,
    sleep_seconds: float = 0.35,
) -> None:
    last_error: Optional[Exception] = None
    for attempt in range(retries + 1):
        try:
            conn.commit()
            return
        except sqlite3.OperationalError as exc:
            last_error = exc
            message = str(exc).lower()
            retriable = "disk i/o error" in message or "database is locked" in message
            if not retriable or attempt >= retries:
                raise
            time.sleep(sleep_seconds * (attempt + 1))
    if last_error:
        raise last_error


def _ensure_orders_operational_columns(conn: sqlite3.Connection) -> bool:
    rows = conn.execute("PRAGMA table_info(orders)").fetchall()
    existing_columns = {str(row["name"]) for row in rows}
    expected_columns = {
        "has_ecommerce_link": "INTEGER NOT NULL DEFAULT 0",
        "has_known_marker": "INTEGER NOT NULL DEFAULT 0",
        "flag_missing_link_and_marker": "INTEGER NOT NULL DEFAULT 0",
        "flag_open_over_2_days": "INTEGER NOT NULL DEFAULT 0",
        "exclude_from_revenue": "INTEGER NOT NULL DEFAULT 0",
        "operational_checked_at": "TEXT",
        "shipping_method_code": "TEXT",
        "shipping_method_label": "TEXT",
    }

    changed = False
    for column_name, column_def in expected_columns.items():
        if column_name in existing_columns:
            continue
        conn.execute(f"ALTER TABLE orders ADD COLUMN {column_name} {column_def}")
        changed = True

    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_orders_flag_missing "
        "ON orders(store_code, flag_missing_link_and_marker)"
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_orders_flag_open_over2 "
        "ON orders(store_code, flag_open_over_2_days)"
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_orders_exclude_revenue "
        "ON orders(store_code, exclude_from_revenue)"
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_orders_shipping_method "
        "ON orders(store_code, shipping_method_label)"
    )

    return changed


def _ensure_operational_watch_columns(conn: sqlite3.Connection) -> bool:
    rows = conn.execute("PRAGMA table_info(operational_watch_orders)").fetchall()
    existing_columns = {str(row["name"]) for row in rows}
    expected_columns = {
        "customer_name": "TEXT",
    }

    changed = False
    for column_name, column_def in expected_columns.items():
        if column_name in existing_columns:
            continue
        conn.execute(f"ALTER TABLE operational_watch_orders ADD COLUMN {column_name} {column_def}")
        changed = True

    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_operational_watch_active "
        "ON operational_watch_orders(active_issue, order_date)"
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_operational_watch_issue_missing "
        "ON operational_watch_orders(issue_missing_link_marker)"
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_operational_watch_issue_open "
        "ON operational_watch_orders(issue_open_over_2_days)"
    )

    return changed


def _enrich_operational_watch_customer_names(conn: sqlite3.Connection) -> int:
    rows = conn.execute(
        "SELECT tiny_id FROM operational_watch_orders "
        "WHERE customer_name IS NULL OR trim(customer_name) = ''"
    ).fetchall()
    if not rows:
        return 0

    updated = 0
    for row in rows:
        tiny_id = row["tiny_id"]
        if tiny_id is None:
            continue
        order_row = conn.execute(
            "SELECT customer_name FROM orders WHERE tiny_id = ? LIMIT 1",
            (int(tiny_id),),
        ).fetchone()
        customer_name = (order_row["customer_name"] if order_row else None) if order_row else None
        if not customer_name:
            continue
        conn.execute(
            "UPDATE operational_watch_orders SET customer_name = ? WHERE tiny_id = ?",
            (str(customer_name).strip(), int(tiny_id)),
        )
        updated += 1
    return updated


def _refresh_operational_flags(conn: sqlite3.Connection, source: Optional[str] = None) -> int:
    sql = (
        "SELECT id, order_date, created_at, situacao, numero_ecommerce, ecommerce_name, markers_text "
        "FROM orders"
    )
    params: tuple[Any, ...] = ()
    if source:
        sql += " WHERE store_code = ?"
        params = (source,)

    rows = conn.execute(sql, params).fetchall()
    if not rows:
        return 0

    checked_at = _now_utc_iso()
    updates: list[tuple[int, int, int, int, int, str, int]] = []
    for row in rows:
        markers = _markers_from_text(row["markers_text"])
        flags = _compute_operational_flags(
            order_date=_parse_date(row["order_date"]),
            created_at=_parse_date(row["created_at"]),
            situacao=row["situacao"],
            numero_ecommerce=row["numero_ecommerce"],
            ecommerce_name=row["ecommerce_name"],
            markers=markers,
        )
        updates.append(
            (
                int(flags["hasEcommerceLink"]),
                int(flags["hasKnownMarker"]),
                int(flags["flagMissingLinkAndMarker"]),
                int(flags["flagOpenOverNDays"]),
                int(_has_excluded_revenue_marker(markers)),
                checked_at,
                int(row["id"]),
            )
        )

    conn.executemany(
        "UPDATE orders SET "
        "has_ecommerce_link = ?, "
        "has_known_marker = ?, "
        "flag_missing_link_and_marker = ?, "
        "flag_open_over_2_days = ?, "
        "exclude_from_revenue = ?, "
        "operational_checked_at = ? "
        "WHERE id = ?",
        updates,
    )
    return len(updates)


def _refresh_operational_watch_issues(conn: sqlite3.Connection) -> int:
    rows = conn.execute(
        # w (watch table) is preferred for flag-relevant fields: the watch loop fetches
        # fresh data from Tiny, so w.situacao / w.markers_text are more current than
        # o.* (orders table), which may be stale for orders outside the lookback window.
        "SELECT "
        "w.tiny_id AS tiny_id, "
        "COALESCE(o.order_date, w.order_date) AS order_date, "
        "COALESCE(o.created_at, w.created_at) AS created_at, "
        "COALESCE(w.situacao, o.situacao) AS situacao, "
        "COALESCE(w.numero_ecommerce, o.numero_ecommerce) AS numero_ecommerce, "
        "COALESCE(w.ecommerce_name, o.ecommerce_name) AS ecommerce_name, "
        "COALESCE(w.markers_text, o.markers_text) AS markers_text, "
        "w.active_issue AS active_issue, "
        "w.resolved_at AS resolved_at "
        "FROM operational_watch_orders w "
        "LEFT JOIN orders o ON o.tiny_id = w.tiny_id"
    ).fetchall()
    if not rows:
        return 0

    now_iso = _now_utc_iso()
    updates: list[tuple[int, int, int, Optional[str], int]] = []
    for row in rows:
        tiny_id = row["tiny_id"]
        if tiny_id is None:
            continue

        flags = _compute_operational_flags(
            order_date=_parse_date(row["order_date"]),
            created_at=_parse_date(row["created_at"]),
            situacao=row["situacao"],
            numero_ecommerce=row["numero_ecommerce"],
            ecommerce_name=row["ecommerce_name"],
            markers=_markers_from_text(row["markers_text"]),
        )
        issue_missing = int(flags["flagMissingLinkAndMarker"])
        issue_open_over_2 = int(flags["flagOpenOverNDays"])
        active_issue = int(issue_missing == 1 or issue_open_over_2 == 1)

        previous_active = int(row["active_issue"] or 0)
        resolved_at = row["resolved_at"]
        if active_issue == 1:
            resolved_value = None
        elif previous_active == 1:
            resolved_value = resolved_at or now_iso
        else:
            resolved_value = resolved_at

        updates.append(
            (
                issue_missing,
                issue_open_over_2,
                active_issue,
                resolved_value,
                int(tiny_id),
            )
        )

    conn.executemany(
        "UPDATE operational_watch_orders SET "
        "issue_missing_link_marker = ?, "
        "issue_open_over_2_days = ?, "
        "active_issue = ?, "
        "resolved_at = ? "
        "WHERE tiny_id = ?",
        updates,
    )
    return len(updates)


def _refresh_shipping_methods_from_raw_payload(conn: sqlite3.Connection) -> int:
    rows = conn.execute("SELECT id, raw_payload FROM orders").fetchall()
    if not rows:
        return 0

    updates: list[tuple[Optional[str], Optional[str], int]] = []
    for row in rows:
        shipping_code = None
        shipping_label = None
        raw_text = row["raw_payload"]
        if raw_text:
            try:
                raw_payload = json.loads(raw_text)
            except ValueError:
                raw_payload = None
            if isinstance(raw_payload, dict):
                shipping_code, shipping_label = _extract_shipping_method(raw_payload)
        updates.append((shipping_code, shipping_label, int(row["id"])))

    conn.executemany(
        "UPDATE orders SET shipping_method_code = ?, shipping_method_label = ? WHERE id = ?",
        updates,
    )
    return len(updates)


def _bootstrap_operational_watch_from_orders(conn: sqlite3.Connection) -> int:
    count_row = conn.execute("SELECT COUNT(*) AS count FROM operational_watch_orders").fetchone()
    if int((count_row["count"] if count_row else 0) or 0) > 0:
        return 0

    rows = conn.execute(
        "SELECT tiny_id, order_number, order_date, created_at, customer_name, situacao, numero_ecommerce, ecommerce_name, markers_text, raw_payload "
        "FROM orders "
        "WHERE tiny_id IS NOT NULL"
    ).fetchall()
    seeded = 0
    for row in rows:
        tiny_id = row["tiny_id"]
        if tiny_id is None:
            continue
        markers = _markers_from_text(row["markers_text"])
        order_payload = {
            "tiny_id": int(tiny_id),
            "number": row["order_number"],
            "date": _parse_date(row["order_date"]),
            "created_at": _parse_date(row["created_at"]),
            "customer_name": row["customer_name"],
            "situacao": row["situacao"],
            "numero_ecommerce": row["numero_ecommerce"],
            "ecommerce_name": row["ecommerce_name"],
            "markers": markers,
        }
        raw_payload = None
        raw_text = row["raw_payload"]
        if raw_text:
            try:
                raw_payload = json.loads(raw_text)
            except Exception:  # noqa: BLE001
                raw_payload = None
        _upsert_operational_watch(conn, order_payload, raw_payload=raw_payload)
        seeded += 1
    return seeded


_REQUIRED_TABLES = {
    "stores",
    "orders",
    "order_items",
    "operational_watch_orders",
    "revenue_targets_monthly",
    "sync_runs",
    "sync_state",
}


def _database_requires_migration(conn: sqlite3.Connection) -> bool:
    table_rows = conn.execute("SELECT name FROM sqlite_master WHERE type = 'table'").fetchall()
    existing_tables = {str(row["name"]) for row in table_rows}
    if not _REQUIRED_TABLES.issubset(existing_tables):
        return True

    orders_columns = {
        str(row["name"])
        for row in conn.execute("PRAGMA table_info(orders)").fetchall()
    }
    expected_orders_columns = {
        "has_ecommerce_link",
        "has_known_marker",
        "flag_missing_link_and_marker",
        "flag_open_over_2_days",
        "exclude_from_revenue",
        "operational_checked_at",
        "shipping_method_code",
        "shipping_method_label",
    }
    if not expected_orders_columns.issubset(orders_columns):
        return True

    watch_columns = {
        str(row["name"])
        for row in conn.execute("PRAGMA table_info(operational_watch_orders)").fetchall()
    }
    if "customer_name" not in watch_columns:
        return True

    return False


def list_sources() -> list[str]:
    return sorted(STORE_RULES.keys())


def _repair_stale_sync_runs(conn: sqlite3.Connection) -> None:
    try:
        conn.execute(
            "UPDATE sync_runs SET status = 'aborted', finished_at = ? "
            "WHERE status = 'running'",
            (_now_utc_iso(),),
        )
    except sqlite3.OperationalError as exc:
        message = str(exc).lower()
        if "no such table" in message or "locked" in message:
            return
        raise


def ensure_database(force: bool = False) -> None:
    global _DB_READY, _STALE_SYNC_RUNS_REPAIRED
    if _DB_READY and _STALE_SYNC_RUNS_REPAIRED and not force:
        return

    with _DB_INIT_LOCK:
        if _DB_READY and _STALE_SYNC_RUNS_REPAIRED and not force:
            return

        DB_PATH.parent.mkdir(parents=True, exist_ok=True)
        conn = _db_connect()
        try:
            if not _STALE_SYNC_RUNS_REPAIRED:
                _repair_stale_sync_runs(conn)
                _safe_commit(conn)
                _STALE_SYNC_RUNS_REPAIRED = True

            if _shipping_labels_need_refresh(conn):
                _refresh_shipping_methods_from_raw_payload(conn)
                _safe_commit(conn)

            # Always ensure GA4 table exists (safe: IF NOT EXISTS)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS ga4_daily_metrics (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    date TEXT NOT NULL,
                    source TEXT NOT NULL DEFAULT 'all',
                    sessions INTEGER,
                    users INTEGER,
                    new_users INTEGER,
                    engaged_sessions INTEGER,
                    add_to_cart INTEGER,
                    checkouts INTEGER,
                    purchases INTEGER,
                    revenue REAL,
                    created_at TEXT NOT NULL DEFAULT (datetime('now'))
                )
            """)
            conn.execute("""
                CREATE UNIQUE INDEX IF NOT EXISTS idx_ga4_daily_source_date
                    ON ga4_daily_metrics(source, date)
            """)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS ad_spend_monthly (
                    store_code TEXT    NOT NULL,
                    year       INTEGER NOT NULL,
                    month      INTEGER NOT NULL,
                    platform   TEXT    NOT NULL,
                    amount     REAL,
                    PRIMARY KEY (store_code, year, month, platform)
                )
            """)
            conn.commit()

            if not force and not _database_requires_migration(conn):
                _DB_READY = True
                return

            conn.execute("PRAGMA journal_mode = WAL")
            conn.execute("PRAGMA synchronous = NORMAL")
            conn.executescript(SCHEMA_PATH.read_text(encoding="utf-8"))
            conn.execute("""
                CREATE TABLE IF NOT EXISTS ga4_daily_metrics (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    date TEXT NOT NULL,
                    source TEXT NOT NULL DEFAULT 'all',
                    sessions INTEGER,
                    users INTEGER,
                    new_users INTEGER,
                    engaged_sessions INTEGER,
                    add_to_cart INTEGER,
                    checkouts INTEGER,
                    purchases INTEGER,
                    revenue REAL,
                    created_at TEXT NOT NULL DEFAULT (datetime('now'))
                )
            """)
            conn.execute("""
                CREATE UNIQUE INDEX IF NOT EXISTS idx_ga4_daily_source_date
                    ON ga4_daily_metrics(source, date)
            """)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS ad_spend_monthly (
                    store_code TEXT    NOT NULL,
                    year       INTEGER NOT NULL,
                    month      INTEGER NOT NULL,
                    platform   TEXT    NOT NULL,
                    amount     REAL,
                    PRIMARY KEY (store_code, year, month, platform)
                )
            """)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS monthly_aggregates (
                    store_code     TEXT    NOT NULL,
                    year           INTEGER NOT NULL,
                    month          INTEGER NOT NULL,
                    total_revenue  REAL    NOT NULL DEFAULT 0.0,
                    order_count    INTEGER NOT NULL DEFAULT 0,
                    avg_ticket     REAL    NOT NULL DEFAULT 0.0,
                    unique_customers INTEGER NOT NULL DEFAULT 0,
                    computed_at    TEXT    NOT NULL,
                    PRIMARY KEY (store_code, year, month)
                )
            """)
            conn.execute("""
                CREATE INDEX IF NOT EXISTS idx_monthly_aggregates_store_year
                ON monthly_aggregates(store_code, year)
            """)
            migration_changed = _ensure_orders_operational_columns(conn)
            watch_migration_changed = _ensure_operational_watch_columns(conn)
            _repair_stale_sync_runs(conn)
            for code, rule in STORE_RULES.items():
                conn.execute(
                    "INSERT INTO stores (code, label, store_type, ecommerce_name, marker_tag, active) "
                    "VALUES (?, ?, ?, ?, ?, 1) "
                    "ON CONFLICT(code) DO UPDATE SET "
                    "label = excluded.label, "
                    "store_type = excluded.store_type, "
                    "ecommerce_name = excluded.ecommerce_name, "
                    "marker_tag = excluded.marker_tag, "
                    "updated_at = CURRENT_TIMESTAMP",
                    (
                        code,
                        rule["label"],
                        rule["type"],
                        rule.get("ecommerce_name"),
                        rule.get("marker_tag"),
                    ),
                )
            _bootstrap_operational_watch_from_orders(conn)
            conn.execute("PRAGMA optimize")
            _enrich_operational_watch_customer_names(conn)
            if migration_changed:
                _refresh_operational_flags(conn, source=None)
                _refresh_shipping_methods_from_raw_payload(conn)
            if migration_changed or watch_migration_changed:
                _refresh_operational_watch_issues(conn)
            _safe_commit(conn)
            _DB_READY = True
        finally:
            conn.close()


def _resolve_spreadsheet_path(path_value: str | Path) -> Path:
    candidate = Path(path_value)
    if candidate.exists():
        return candidate
    local_candidate = (TINY_DIR / candidate).resolve()
    if local_candidate.exists():
        return local_candidate
    return candidate


def import_brew_monthly_history_workbook(
    path: str | Path = DEFAULT_BREW_HISTORY_FILE,
    *,
    start_year: int = 2021,
    end_year: int = 2024,
) -> dict[str, Any]:
    if end_year < start_year:
        raise ValueError("Intervalo de anos invalido.")

    resolved_path = _resolve_spreadsheet_path(path)
    if not resolved_path.exists():
        raise ValueError(f"Arquivo de historico nao encontrado: {resolved_path}")

    ensure_database()

    workbook = load_workbook(resolved_path, data_only=True, read_only=True)
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        selected_sheet = worksheet.title

        year_columns: dict[int, int] = {}
        parsed_rows: dict[tuple[int, int], float] = {}
        month_rows = 0

        for row in worksheet.iter_rows(min_row=1, values_only=True):
            if not row:
                continue

            for index, cell_value in enumerate(row):
                year_value = _parse_int(cell_value)
                if year_value is not None and 2000 <= year_value <= 2100:
                    year_columns[index] = year_value

            month = None
            for cell_value in row:
                month = _month_from_header(cell_value)
                if month is not None:
                    break
            if month is None:
                continue

            month_rows += 1
            for col_index, year_value in year_columns.items():
                if year_value < start_year or year_value > end_year:
                    continue
                if col_index >= len(row):
                    continue
                revenue_value = _parse_float(row[col_index])
                if revenue_value is None:
                    continue
                parsed_rows[(year_value, month)] = revenue_value

        if not year_columns:
            raise ValueError("Planilha sem cabecalho de anos.")
        if not parsed_rows:
            raise ValueError("Planilha sem valores de faturamento no intervalo informado.")

        conn = _db_connect()
        upserted = 0
        overwritten = 0
        inserted = 0
        preserved_targets = 0
        try:
            for (year, month), realized_revenue in sorted(parsed_rows.items()):
                existing_row = conn.execute(
                    """
                    SELECT target_revenue, realized_revenue
                    FROM revenue_targets_monthly
                    WHERE store_code = ? AND year = ? AND month = ?
                    """,
                    ("brew", year, month),
                ).fetchone()

                if existing_row:
                    current_realized = (
                        float(existing_row["realized_revenue"])
                        if existing_row["realized_revenue"] is not None
                        else None
                    )
                    if current_realized != realized_revenue:
                        overwritten += 1
                    if existing_row["target_revenue"] is not None:
                        preserved_targets += 1
                else:
                    inserted += 1

                conn.execute(
                    """
                    INSERT INTO revenue_targets_monthly
                        (store_code, year, month, target_revenue, realized_revenue, source_file)
                    VALUES (?, ?, ?, ?, ?, ?)
                    ON CONFLICT(store_code, year, month) DO UPDATE SET
                        target_revenue = revenue_targets_monthly.target_revenue,
                        realized_revenue = excluded.realized_revenue,
                        source_file = excluded.source_file,
                        loaded_at = CURRENT_TIMESTAMP
                    """,
                    ("brew", year, month, None, realized_revenue, resolved_path.name),
                )
                upserted += 1
            _safe_commit(conn)
        finally:
            conn.close()
    finally:
        workbook.close()

    imported_years = sorted({year for year, _ in parsed_rows})
    return {
        "status": "success" if upserted else "failed",
        "source": "brew",
        "file": str(resolved_path),
        "sheet": selected_sheet,
        "rowsUpserted": upserted,
        "rowsInserted": inserted,
        "rowsOverwritten": overwritten,
        "targetsPreserved": preserved_targets,
        "monthRowsFound": month_rows,
        "yearsImported": imported_years,
        "yearRange": [start_year, end_year],
    }


def import_revenue_targets_workbook(
    path: Path,
    source: str,
    sheet_name: str | None = "Metas_e_Realizado",
) -> dict[str, Any]:
    if source not in {"brew", "grow"}:
        raise ValueError("Importacao de metas habilitada apenas para e-commerces (brew/grow).")

    resolved_path = _resolve_spreadsheet_path(path)
    if not resolved_path.exists():
        raise ValueError(f"Arquivo de metas nao encontrado: {resolved_path}")

    ensure_database()

    workbook = load_workbook(resolved_path, data_only=True, read_only=True)
    try:
        selected_sheet = sheet_name if sheet_name and sheet_name in workbook.sheetnames else workbook.sheetnames[0]
        worksheet = workbook[selected_sheet]

        rows_iter = worksheet.iter_rows(min_row=1, values_only=True)
        header = next(rows_iter, None)
        if not header:
            raise ValueError("Planilha sem cabecalho.")

        month_columns: dict[int, int] = {}
        for index, cell_value in enumerate(header):
            month = _month_from_header(cell_value)
            if month is not None:
                month_columns[index] = month
        if not month_columns:
            raise ValueError("Planilha sem colunas mensais JAN..DEZ.")

        parsed_rows: dict[tuple[int, int], dict[str, Optional[float]]] = {}
        invalid_rows = 0
        target_cells = 0
        realized_cells = 0

        for row in rows_iter:
            if not row:
                continue
            year = _parse_int(row[0] if len(row) > 0 else None)
            row_type = _target_row_type(row[1] if len(row) > 1 else None)
            if year is None or row_type is None:
                if any(cell not in (None, "") for cell in row[:2]):
                    invalid_rows += 1
                continue

            for col_index, month in month_columns.items():
                raw_value = row[col_index] if col_index < len(row) else None
                value = _parse_float(raw_value)
                if value is None:
                    continue
                key = (year, month)
                entry = parsed_rows.get(key)
                if not entry:
                    entry = {"target": None, "realized": None}
                    parsed_rows[key] = entry
                entry[row_type] = value
                if row_type == "target":
                    target_cells += 1
                else:
                    realized_cells += 1

        conn = _db_connect()
        upserted = 0
        try:
            for (year, month), values in sorted(parsed_rows.items()):
                if values.get("target") is None and values.get("realized") is None:
                    continue
                conn.execute(
                    "INSERT INTO revenue_targets_monthly "
                    "(store_code, year, month, target_revenue, realized_revenue, source_file) "
                    "VALUES (?, ?, ?, ?, ?, ?) "
                    "ON CONFLICT(store_code, year, month) DO UPDATE SET "
                    "target_revenue = COALESCE(excluded.target_revenue, revenue_targets_monthly.target_revenue), "
                    "realized_revenue = COALESCE(excluded.realized_revenue, revenue_targets_monthly.realized_revenue), "
                    "source_file = excluded.source_file, "
                    "loaded_at = CURRENT_TIMESTAMP",
                    (
                        source,
                        year,
                        month,
                        values.get("target"),
                        values.get("realized"),
                        resolved_path.name,
                    ),
                )
                upserted += 1
            _safe_commit(conn)
        finally:
            conn.close()
    finally:
        workbook.close()

    return {
        "status": "success" if upserted else "failed",
        "source": source,
        "file": str(resolved_path),
        "sheet": selected_sheet,
        "rowsUpserted": upserted,
        "targetCells": target_cells,
        "realizedCells": realized_cells,
        "invalidRows": invalid_rows,
    }


def load_default_ecommerce_targets(sheet_name: str = "Metas_e_Realizado") -> dict[str, Any]:
    results: list[dict[str, Any]] = []
    for source, default_path in DEFAULT_ECOM_TARGET_FILES.items():
        env_key = f"TINY_TARGETS_FILE_{source.upper()}"
        custom_path = os.getenv(env_key)
        source_path = _resolve_spreadsheet_path(custom_path) if custom_path else default_path
        if not source_path.exists():
            results.append(
                {
                    "status": "skipped",
                    "source": source,
                    "file": str(source_path),
                    "reason": "arquivo_nao_encontrado",
                }
            )
            continue

        try:
            result = import_revenue_targets_workbook(source_path, source, sheet_name=sheet_name)
        except Exception as exc:  # noqa: BLE001
            result = {
                "status": "failed",
                "source": source,
                "file": str(source_path),
                "error": str(exc),
            }
        results.append(result)

    success = sum(1 for item in results if item.get("status") == "success")
    return {
        "mode": "load_default_ecommerce_targets",
        "successCount": success,
        "results": results,
    }


def _create_sync_run(
    conn: sqlite3.Connection,
    source: str,
    mode: str,
    requested_from: Optional[date],
    requested_to: Optional[date],
) -> int:
    now = _now_utc_iso()
    cur = conn.execute(
        "INSERT INTO sync_runs "
        "(store_code, mode, requested_from, requested_to, started_at, status) "
        "VALUES (?, ?, ?, ?, ?, 'running')",
        (
            source,
            mode,
            requested_from.isoformat() if requested_from else None,
            requested_to.isoformat() if requested_to else None,
            now,
        ),
    )
    return int(cur.lastrowid)


def _finalize_sync_run(
    conn: sqlite3.Connection,
    run_id: int,
    source: str,
    status: str,
    orders_found: int,
    orders_synced: int,
    errors_count: int,
    details: dict[str, Any],
    last_order_date: Optional[date],
) -> None:
    finished = _now_utc_iso()
    conn.execute(
        "UPDATE sync_runs SET "
        "status = ?, finished_at = ?, orders_found = ?, orders_synced = ?, errors_count = ?, details = ? "
        "WHERE id = ?",
        (status, finished, orders_found, orders_synced, errors_count, json.dumps(details, ensure_ascii=False), run_id),
    )

    if status in {"success", "partial"}:
        conn.execute(
            "INSERT INTO sync_state (store_code, last_success_at, last_order_date, last_run_id, updated_at) "
            "VALUES (?, ?, ?, ?, CURRENT_TIMESTAMP) "
            "ON CONFLICT(store_code) DO UPDATE SET "
            "last_success_at = excluded.last_success_at, "
            "last_order_date = COALESCE(excluded.last_order_date, sync_state.last_order_date), "
            "last_run_id = excluded.last_run_id, "
            "updated_at = CURRENT_TIMESTAMP",
            (
                source,
                finished,
                last_order_date.isoformat() if last_order_date else None,
                run_id,
            ),
        )


def _find_existing_order_id(
    conn: sqlite3.Connection,
    source: str,
    tiny_id: Optional[int],
    order_number: str,
    order_date: date,
) -> Optional[int]:
    if tiny_id is not None:
        row = conn.execute("SELECT id FROM orders WHERE tiny_id = ?", (tiny_id,)).fetchone()
        if row:
            return int(row["id"])

    row = conn.execute(
        "SELECT id FROM orders WHERE store_code = ? AND order_number = ? AND order_date = ?",
        (source, order_number, order_date.isoformat()),
    ).fetchone()
    return int(row["id"]) if row else None


def _existing_tiny_ids(conn: sqlite3.Connection, ids: list[int]) -> set[int]:
    if not ids:
        return set()
    existing: set[int] = set()
    chunk_size = 900
    for index in range(0, len(ids), chunk_size):
        chunk = ids[index : index + chunk_size]
        placeholders = ",".join(["?"] * len(chunk))
        rows = conn.execute(
            f"SELECT tiny_id FROM orders WHERE tiny_id IN ({placeholders})",
            chunk,
        ).fetchall()
        for row in rows:
            if row["tiny_id"] is not None:
                existing.add(int(row["tiny_id"]))
    return existing


def _existing_order_snapshots(
    conn: sqlite3.Connection,
    source: str,
    ids: list[int],
) -> dict[int, dict[str, Any]]:
    if not ids:
        return {}
    snapshots: dict[int, dict[str, Any]] = {}
    chunk_size = 900
    for index in range(0, len(ids), chunk_size):
        chunk = ids[index : index + chunk_size]
        placeholders = ",".join(["?"] * len(chunk))
        rows = conn.execute(
            f"SELECT tiny_id, situacao, total FROM orders WHERE store_code = ? AND tiny_id IN ({placeholders})",
            [source, *chunk],
        ).fetchall()
        for row in rows:
            tiny_id = row["tiny_id"]
            if tiny_id is None:
                continue
            snapshots[int(tiny_id)] = {
                "situacao": str(row["situacao"] or "").strip() or None,
                "total": float(row["total"]) if row["total"] is not None else None,
            }
    return snapshots


def _extract_summary_situacao(summary: dict[str, Any]) -> Optional[str]:
    situacao_payload = summary.get("situacao")
    if isinstance(situacao_payload, dict):
        value = (
            situacao_payload.get("descricao")
            or situacao_payload.get("situacao")
            or situacao_payload.get("nome")
        )
    else:
        value = situacao_payload
    text = str(value or "").strip()
    return text or None


def _extract_summary_total(summary: dict[str, Any]) -> Optional[float]:
    for key in ("valor", "total_pedido", "total"):
        parsed = _parse_float(summary.get(key))
        if parsed is not None:
            return parsed
    return None


def _summary_indicates_order_change(
    summary: dict[str, Any],
    snapshot: dict[str, Any],
    *,
    total_tolerance: float = 0.009,
) -> bool:
    summary_situacao = _extract_summary_situacao(summary)
    db_situacao = snapshot.get("situacao")

    if _normalize_token(summary_situacao) != _normalize_token(db_situacao):
        return True

    summary_total = _extract_summary_total(summary)
    db_total_raw = snapshot.get("total")
    db_total = float(db_total_raw) if db_total_raw is not None else None

    if summary_total is None and db_total is None:
        return False
    if summary_total is None or db_total is None:
        return True
    return abs(summary_total - db_total) > total_tolerance


def _existing_watch_tiny_ids(conn: sqlite3.Connection, ids: list[int]) -> set[int]:
    if not ids:
        return set()
    existing: set[int] = set()
    chunk_size = 900
    for index in range(0, len(ids), chunk_size):
        chunk = ids[index : index + chunk_size]
        placeholders = ",".join(["?"] * len(chunk))
        rows = conn.execute(
            f"SELECT tiny_id FROM operational_watch_orders WHERE tiny_id IN ({placeholders})",
            chunk,
        ).fetchall()
        for row in rows:
            if row["tiny_id"] is not None:
                existing.add(int(row["tiny_id"]))
    return existing


def _operational_watch_tiny_ids(
    conn: sqlite3.Connection,
    *,
    limit: int = OPERATIONAL_WATCH_MAX_ORDERS,
) -> list[int]:
    rows = conn.execute(
        "SELECT tiny_id "
        "FROM operational_watch_orders "
        "WHERE tiny_id IS NOT NULL "
        "AND active_issue = 1 "
        "ORDER BY date(order_date) DESC, tiny_id DESC "
        "LIMIT ?",
        (max(1, int(limit)),),
    ).fetchall()
    return [int(row["tiny_id"]) for row in rows if row["tiny_id"] is not None]


def _upsert_operational_watch(
    conn: sqlite3.Connection,
    order: dict[str, Any],
    *,
    raw_payload: Optional[dict[str, Any]],
) -> None:
    tiny_id_value = order.get("tiny_id")
    tiny_id = int(tiny_id_value) if tiny_id_value is not None and str(tiny_id_value).isdigit() else None
    if tiny_id is None:
        return

    order_date_value = order.get("date")
    created_at_value = order.get("created_at")
    order_date = order_date_value.isoformat() if isinstance(order_date_value, date) else None
    created_at = created_at_value.isoformat() if isinstance(created_at_value, date) else None
    markers = order.get("markers") or []
    markers_text = ",".join(str(m) for m in markers if str(m).strip())
    flags = _compute_operational_flags(
        order_date=order_date_value if isinstance(order_date_value, date) else None,
        created_at=created_at_value if isinstance(created_at_value, date) else None,
        situacao=order.get("situacao"),
        numero_ecommerce=order.get("numero_ecommerce"),
        ecommerce_name=order.get("ecommerce_name"),
        markers=markers,
    )
    issue_missing = int(flags["flagMissingLinkAndMarker"])
    issue_open_over_2 = int(flags["flagOpenOverNDays"])
    active_issue = int(issue_missing == 1 or issue_open_over_2 == 1)
    payload_json = json.dumps(raw_payload, ensure_ascii=False) if raw_payload is not None else None
    now_iso = _now_utc_iso()
    customer_name = str(order.get("customer_name") or "").strip() or None

    conn.execute(
        "INSERT INTO operational_watch_orders ("
        "tiny_id, order_number, order_date, created_at, customer_name, situacao, numero_ecommerce, ecommerce_name, markers_text, "
        "issue_missing_link_marker, issue_open_over_2_days, active_issue, first_seen_at, last_seen_at, resolved_at, raw_payload"
        ") VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?) "
        "ON CONFLICT(tiny_id) DO UPDATE SET "
        "order_number = excluded.order_number, "
        "order_date = excluded.order_date, "
        "created_at = excluded.created_at, "
        "customer_name = COALESCE(excluded.customer_name, operational_watch_orders.customer_name), "
        "situacao = excluded.situacao, "
        "numero_ecommerce = excluded.numero_ecommerce, "
        "ecommerce_name = excluded.ecommerce_name, "
        "markers_text = excluded.markers_text, "
        "issue_missing_link_marker = excluded.issue_missing_link_marker, "
        "issue_open_over_2_days = excluded.issue_open_over_2_days, "
        "active_issue = excluded.active_issue, "
        "last_seen_at = excluded.last_seen_at, "
        "resolved_at = CASE "
        "  WHEN excluded.active_issue = 1 THEN NULL "
        "  WHEN operational_watch_orders.active_issue = 1 THEN excluded.last_seen_at "
        "  ELSE operational_watch_orders.resolved_at "
        "END, "
        "raw_payload = COALESCE(excluded.raw_payload, operational_watch_orders.raw_payload)",
        (
            tiny_id,
            str(order.get("number") or ""),
            order_date,
            created_at,
            customer_name,
            order.get("situacao"),
            str(order.get("numero_ecommerce") or "") or None,
            str(order.get("ecommerce_name") or "") or None,
            markers_text,
            issue_missing,
            issue_open_over_2,
            active_issue,
            now_iso,
            now_iso,
            None if active_issue else now_iso,
            payload_json,
        ),
    )


def _upsert_normalized_order(
    conn: sqlite3.Connection,
    source: str,
    order: dict[str, Any],
    *,
    raw_payload: Optional[dict[str, Any]],
) -> int:
    order_number = str(order.get("number") or "").strip()
    order_date = order.get("date")

    if not order_number:
        raise ValueError("Pedido sem numero.")
    if not isinstance(order_date, date):
        raise ValueError("Pedido sem data valida.")

    tiny_id_value = order.get("tiny_id")
    tiny_id = int(tiny_id_value) if tiny_id_value is not None and str(tiny_id_value).isdigit() else None
    existing_id = _find_existing_order_id(conn, source, tiny_id, order_number, order_date)

    markers = order.get("markers") or []
    markers_text = ",".join(str(m) for m in markers)
    exclude_from_revenue = int(_has_excluded_revenue_marker(markers))

    payload_json = json.dumps(raw_payload, ensure_ascii=False) if raw_payload is not None else None

    values = (
        tiny_id,
        source,
        order_number,
        order.get("numero_ecommerce"),
        order.get("ecommerce_name"),
        order_date.isoformat(),
        order.get("created_at").isoformat() if isinstance(order.get("created_at"), date) else None,
        order.get("updated_at").isoformat() if isinstance(order.get("updated_at"), date) else None,
        order.get("situacao_id"),
        order.get("situacao"),
        order.get("total"),
        order.get("subtotal"),
        order.get("shipping"),
        order.get("discount"),
        markers_text,
        exclude_from_revenue,
        order.get("payment_method"),
        order.get("shipping_method_code"),
        order.get("shipping_method_label"),
        order.get("customer_code"),
        order.get("customer_name"),
        order.get("customer_email"),
        order.get("customer_document"),
        order.get("customer_city"),
        order.get("customer_state"),
        payload_json,
    )

    if existing_id is None:
        cur = conn.execute(
            "INSERT INTO orders ("
            "tiny_id, store_code, order_number, numero_ecommerce, ecommerce_name, order_date, created_at, "
            "updated_at, situacao_id, situacao, total, subtotal, shipping, discount, markers_text, exclude_from_revenue, payment_method, "
            "shipping_method_code, shipping_method_label, customer_code, customer_name, customer_email, customer_document, customer_city, customer_state, "
            "raw_payload) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            values,
        )
        order_id = int(cur.lastrowid)
    else:
        conn.execute(
            "UPDATE orders SET "
            "tiny_id = ?, store_code = ?, order_number = ?, numero_ecommerce = ?, ecommerce_name = ?, order_date = ?, "
            "created_at = ?, updated_at = ?, situacao_id = ?, situacao = ?, total = ?, subtotal = ?, shipping = ?, "
            "discount = ?, markers_text = ?, exclude_from_revenue = ?, payment_method = ?, shipping_method_code = ?, shipping_method_label = ?, customer_code = ?, customer_name = ?, "
            "customer_email = ?, customer_document = ?, customer_city = ?, customer_state = ?, "
            "raw_payload = ?, synced_at = CURRENT_TIMESTAMP "
            "WHERE id = ?",
            values + (existing_id,),
        )
        order_id = existing_id

    conn.execute("DELETE FROM order_items WHERE order_id = ?", (order_id,))

    for index, item in enumerate(order.get("items") or [], start=1):
        conn.execute(
            "INSERT INTO order_items ("
            "order_id, line_number, sku, product_id, product_name, quantity, unit_price, revenue, raw_payload"
            ") VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (
                order_id,
                int(item.get("line_number") or index),
                item.get("sku"),
                str(item.get("product_id")) if item.get("product_id") is not None else None,
                item.get("name"),
                item.get("quantity"),
                item.get("unit_price"),
                item.get("revenue"),
                json.dumps(item.get("raw"), ensure_ascii=False) if item.get("raw") is not None else None,
            ),
        )

    return order_id


def incremental_window(source: str, lookback_days: int = SYNC_LOOKBACK_DAYS) -> tuple[date, date]:
    ensure_database()
    conn = _db_connect()
    try:
        row = conn.execute(
            "SELECT last_order_date FROM sync_state WHERE store_code = ?",
            (source,),
        ).fetchone()

        anchor = None
        if row and row["last_order_date"]:
            anchor = _parse_date(row["last_order_date"])

        if anchor is None:
            row = conn.execute(
                "SELECT MAX(order_date) AS max_date FROM orders WHERE store_code = ?",
                (source,),
            ).fetchone()
            if row and row["max_date"]:
                anchor = _parse_date(row["max_date"])

        today = date.today()
        if anchor is None:
            start = date(today.year, 1, 1)
        else:
            start = anchor - timedelta(days=max(1, lookback_days))

        if start > today:
            start = today

        return start, today
    finally:
        conn.close()


def _incremental_audit_from(end: date, audit_days: int = INCREMENTAL_AUDIT_DAYS) -> Optional[date]:
    if audit_days <= 0:
        return None
    return end - timedelta(days=max(1, audit_days))


def sync_store(
    source: str,
    start: date,
    end: date,
    *,
    mode: str,
    detail_workers: int = DETAIL_MAX_WORKERS,
    capture_operational_scan: bool = False,
    audit_from: Optional[date] = None,
) -> dict[str, Any]:
    if source not in STORE_RULES:
        raise ValueError(f"Fonte desconhecida: {source}")
    if end < start:
        raise ValueError("Intervalo invalido: data final menor que data inicial.")

    ensure_database()

    conn = _db_connect()
    run_id = _create_sync_run(conn, source, mode, start, end)
    _safe_commit(conn)

    fetch_start = start
    if mode == "incremental" and isinstance(audit_from, date) and audit_from <= end:
        fetch_start = min(start, audit_from)

    summaries, error = _fetch_order_summaries(fetch_start, end, source)
    if error:
        _finalize_sync_run(
            conn,
            run_id,
            source,
            "failed",
            0,
            0,
            1,
            {"error": error},
            None,
        )
        _safe_commit(conn)
        conn.close()
        return {
            "runId": run_id,
            "source": source,
            "mode": mode,
            "status": "failed",
            "ordersFound": 0,
            "ordersSynced": 0,
            "errorsCount": 1,
            "error": error,
        }

    filtered_summaries: list[dict[str, Any]] = []
    for summary in summaries:
        order_id = summary.get("id")
        if order_id is None:
            continue
        if (
            STORE_RULES[source].get("type") == "ecommerce"
            and not summary.get("numero_ecommerce")
            and not capture_operational_scan
        ):
            logger.warning(
                "Pedido ecommerce descartado (sem numero_ecommerce): "
                "id=%s situacao=%s data=%s loja=%s",
                summary.get("id"),
                summary.get("situacao"),
                summary.get("data_pedido"),
                source,
            )
            continue
        filtered_summaries.append(summary)

    errors: list[str] = []
    orders_synced = 0
    orders_matched = 0
    latest_date: Optional[date] = None
    operational_watch_candidates = 0
    operational_watch_refreshed = 0
    operational_rows_updated = 0
    operational_watch_rows_updated = 0

    summaries_to_fetch = filtered_summaries
    skipped_existing = 0
    skipped_watch_existing = 0
    changed_existing = 0
    change_candidates = 0
    if mode == "incremental":
        ids = [int(summary["id"]) for summary in filtered_summaries if str(summary.get("id")).isdigit()]
        snapshots = _existing_order_snapshots(conn, source, ids)
        existing_ids = set(snapshots.keys())
        watch_existing_ids = _existing_watch_tiny_ids(conn, ids) if capture_operational_scan else set()
        summaries_to_fetch = []
        for summary in filtered_summaries:
            summary_id = int(summary["id"])
            if summary_id in existing_ids:
                change_candidates += 1
                if _summary_indicates_order_change(summary, snapshots.get(summary_id, {})):
                    changed_existing += 1
                    summaries_to_fetch.append(summary)
                    continue
                skipped_existing += 1
                continue
            if capture_operational_scan and summary_id in watch_existing_ids:
                skipped_watch_existing += 1
                continue
            summaries_to_fetch.append(summary)

    def load_detail(summary: dict[str, Any]) -> tuple[dict[str, Any], Optional[dict[str, Any]], Optional[str]]:
        order_id = int(summary["id"])
        detail, detail_error = _fetch_order_detail_with_retry(order_id)
        return summary, detail, detail_error

    with ThreadPoolExecutor(max_workers=max(1, detail_workers)) as executor:
        futures = [executor.submit(load_detail, summary) for summary in summaries_to_fetch]
        for future in as_completed(futures):
            summary, detail, detail_error = future.result()
            if detail_error or not detail:
                errors.append(
                    f"Pedido {summary.get('id')}: {detail_error or 'sem detalhes'}"
                )
                continue

            normalized = _normalize_order(detail)
            if capture_operational_scan:
                try:
                    _upsert_operational_watch(
                        conn,
                        normalized,
                        raw_payload=detail,
                    )
                except Exception as exc:  # noqa: BLE001
                    errors.append(f"Pedido {summary.get('id')}: falha no rastreamento operacional ({exc})")
            if not _store_match(normalized, source):
                continue

            if not normalized.get("date"):
                errors.append(f"Pedido {summary.get('id')}: sem data valida")
                continue

            orders_matched += 1
            try:
                _upsert_normalized_order(
                    conn,
                    source,
                    normalized,
                    raw_payload=detail,
                )
                orders_synced += 1
                order_date = normalized.get("date")
                if isinstance(order_date, date) and (latest_date is None or order_date > latest_date):
                    latest_date = order_date
                if orders_synced % 100 == 0:
                    _safe_commit(conn)
            except Exception as exc:  # noqa: BLE001
                errors.append(f"Pedido {summary.get('id')}: {exc}")

    if mode == "incremental" and capture_operational_scan:
        already_fetched_ids = {
            int(summary["id"])
            for summary in summaries_to_fetch
            if str(summary.get("id")).isdigit()
        }
        watch_ids = [
            tiny_id
            for tiny_id in _operational_watch_tiny_ids(conn)
            if tiny_id not in already_fetched_ids
        ]
        operational_watch_candidates = len(watch_ids)

        for tiny_id in watch_ids:
            detail, detail_error = _fetch_order_detail_with_retry(int(tiny_id))
            if detail_error or not detail:
                errors.append(f"Pedido {tiny_id}: {detail_error or 'sem detalhes'}")
                continue

            normalized = _normalize_order(detail)
            try:
                _upsert_operational_watch(
                    conn,
                    normalized,
                    raw_payload=detail,
                )
            except Exception as exc:  # noqa: BLE001
                errors.append(f"Pedido {tiny_id}: falha no rastreamento operacional ({exc})")
            # Always propagate the fresh situacao back to the orders table so that
            # _refresh_operational_watch_issues sees current data even for orders
            # outside the incremental lookback window or belonging to another source.
            new_situacao = normalized.get("situacao")
            if new_situacao is not None:
                conn.execute(
                    "UPDATE orders SET situacao = ?, synced_at = CURRENT_TIMESTAMP "
                    "WHERE tiny_id = ? AND situacao != ?",
                    (new_situacao, tiny_id, new_situacao),
                )
            if not _store_match(normalized, source):
                continue
            if not normalized.get("date"):
                errors.append(f"Pedido {tiny_id}: sem data valida")
                continue

            try:
                _upsert_normalized_order(
                    conn,
                    source,
                    normalized,
                    raw_payload=detail,
                )
                operational_watch_refreshed += 1
                order_date = normalized.get("date")
                if isinstance(order_date, date) and (latest_date is None or order_date > latest_date):
                    latest_date = order_date
            except Exception as exc:  # noqa: BLE001
                errors.append(f"Pedido {tiny_id}: {exc}")

    operational_rows_updated = _refresh_operational_flags(conn, source=source)
    operational_watch_rows_updated = _refresh_operational_watch_issues(conn)
    _safe_commit(conn)

    status = "success"
    if errors and orders_synced > 0:
        status = "partial"
    elif errors and orders_synced == 0:
        status = "failed"

    details = {
        "requestedFrom": start.isoformat(),
        "fetchFrom": fetch_start.isoformat(),
        "auditFrom": audit_from.isoformat() if isinstance(audit_from, date) else None,
        "summaries": len(filtered_summaries),
        "queuedForDetails": len(summaries_to_fetch),
        "skippedExisting": skipped_existing,
        "changeCandidates": change_candidates,
        "changedExisting": changed_existing,
        "skippedWatchExisting": skipped_watch_existing,
        "operationalWatchCandidates": operational_watch_candidates,
        "operationalWatchRefreshed": operational_watch_refreshed,
        "matched": orders_matched,
        "operationalRowsUpdated": operational_rows_updated,
        "operationalWatchRowsUpdated": operational_watch_rows_updated,
        "errorsPreview": errors[:20],
    }

    _finalize_sync_run(
        conn,
        run_id,
        source,
        status,
        len(filtered_summaries),
        orders_synced,
        len(errors),
        details,
        latest_date,
    )
    _safe_commit(conn)
    conn.close()

    return {
        "runId": run_id,
        "source": source,
        "mode": mode,
        "status": status,
        "ordersFound": len(filtered_summaries),
        "ordersQueued": len(summaries_to_fetch),
        "ordersSkippedExisting": skipped_existing,
        "ordersMatched": orders_matched,
        "ordersSynced": orders_synced,
        "errorsCount": len(errors),
        "errorsPreview": errors[:20],
        "startDate": start.isoformat(),
        "fetchStartDate": fetch_start.isoformat(),
        "endDate": end.isoformat(),
    }


def _resolve_sources(source: str) -> list[str]:
    if source == "all":
        return list_sources()
    if source not in STORE_RULES:
        raise ValueError(f"Fonte desconhecida: {source}")
    return [source]


def sync_incremental(source: str = "all", lookback_days: int = SYNC_LOOKBACK_DAYS) -> dict[str, Any]:
    results: list[dict[str, Any]] = []
    sources = _resolve_sources(source)
    for src in sources:
        start, end = incremental_window(src, lookback_days)
        audit_from = _incremental_audit_from(end)
        capture_operational_scan = False
        if source == "all":
            capture_operational_scan = src == "brew"
        else:
            capture_operational_scan = src in {"brew", "grow"}
        results.append(
            sync_store(
                src,
                start,
                end,
                mode="incremental",
                capture_operational_scan=capture_operational_scan,
                audit_from=audit_from,
            )
        )
    return {
        "mode": "incremental",
        "source": source,
        "lookbackDays": lookback_days,
        "results": results,
    }


def sync_date_range(source: str, start: date, end: date) -> dict[str, Any]:
    results: list[dict[str, Any]] = []
    for src in _resolve_sources(source):
        capture_operational_scan = (src == "brew") if source == "all" else (src in {"brew", "grow"})
        results.append(
            sync_store(
                src,
                start,
                end,
                mode="manual",
                capture_operational_scan=capture_operational_scan,
            )
        )
    return {
        "mode": "manual",
        "source": source,
        "startDate": start.isoformat(),
        "endDate": end.isoformat(),
        "results": results,
    }


def backfill_years(source: str, years: list[int]) -> dict[str, Any]:
    today = date.today()
    runs: list[dict[str, Any]] = []

    for src in _resolve_sources(source):
        for year in sorted(set(years)):
            for month in range(1, 13):
                start = date(year, month, 1)
                end = date(year, month, calendar.monthrange(year, month)[1])
                if start > today:
                    continue
                if end > today:
                    end = today
                capture_operational_scan = (src == "brew") if source == "all" else (src in {"brew", "grow"})
                runs.append(
                    sync_store(
                        src,
                        start,
                        end,
                        mode="backfill",
                        capture_operational_scan=capture_operational_scan,
                    )
                )

    return {
        "mode": "backfill",
        "source": source,
        "years": sorted(set(years)),
        "runs": runs,
    }


def _daily_totals_for_source(source: str, start: date, end: date) -> dict[date, float]:
    conn = _db_connect()
    try:
        rows = conn.execute(
            "SELECT order_date, SUM(total) AS revenue "
            "FROM orders "
            f"WHERE store_code = ? AND date(order_date) BETWEEN ? AND ? AND {_orders_analytics_filter_sql()} "
            "GROUP BY order_date",
            (source, start.isoformat(), end.isoformat()),
        ).fetchall()
    finally:
        conn.close()

    totals: dict[date, float] = {}
    for row in rows:
        order_date = _parse_date(row["order_date"])
        if not order_date:
            continue
        totals[order_date] = float(row["revenue"] or 0.0)
    return totals


def _build_weekly_entries(
    years: list[int],
    daily_totals: dict[date, float],
    monthly_targets: dict[tuple[int, int], float],
) -> list[dict[str, Any]]:
    entries: list[dict[str, Any]] = []
    for year in years:
        for month in range(1, 13):
            segments = _week_segments(year, month)
            month_days_count = calendar.monthrange(year, month)[1]
            month_target = monthly_targets.get((year, month))
            month_days = [d for d in daily_totals if d.year == year and d.month == month]
            has_month_data = bool(month_days)
            for week_index, start, end in segments:
                if has_month_data:
                    realized = sum(value for day, value in daily_totals.items() if start <= day <= end)
                else:
                    realized = None
                if month_target is not None and month_days_count:
                    segment_days = (end - start).days + 1
                    weekly_target = month_target * (segment_days / month_days_count)
                else:
                    weekly_target = None
                entries.append(
                    {
                        "year": year,
                        "month": month,
                        "week": week_index,
                        "start": start.isoformat(),
                        "end": end.isoformat(),
                        "meta": weekly_target,
                        "realizado": realized,
                    }
                )
    return entries


def build_dataset(source: str, start_year: int, end_year: int) -> dict[str, Any]:
    if source not in STORE_RULES:
        raise ValueError("Fonte desconhecida.")

    ensure_database()

    start = date(start_year, 1, 1)
    end = min(date(end_year, 12, 31), date.today())

    conn = _db_connect()
    try:
        order_rows = conn.execute(
            "SELECT "
            "CAST(strftime('%Y', order_date) AS INTEGER) AS year, "
            "CAST(strftime('%m', order_date) AS INTEGER) AS month, "
            "SUM(total) AS revenue, "
            "COUNT(DISTINCT id) AS orders "
            "FROM orders "
            f"WHERE store_code = ? AND date(order_date) BETWEEN ? AND ? AND {_orders_analytics_filter_sql()} "
            "GROUP BY year, month",
            (source, start.isoformat(), end.isoformat()),
        ).fetchall()

        month_data: dict[tuple[int, int], dict[str, float | int]] = {}
        for row in order_rows:
            year = int(row["year"])
            month = int(row["month"])
            month_data[(year, month)] = {
                "revenue": float(row["revenue"] or 0.0),
                "orders": int(row["orders"] or 0),
            }

        target_rows = conn.execute(
            "SELECT year, month, target_revenue, realized_revenue "
            "FROM revenue_targets_monthly "
            "WHERE store_code = ? AND (year * 100 + month) BETWEEN ? AND ?",
            (source, start_year * 100 + 1, end_year * 100 + 12),
        ).fetchall()
        target_data: dict[tuple[int, int], dict[str, Optional[float]]] = {}
        for row in target_rows:
            key = (int(row["year"]), int(row["month"]))
            target_data[key] = {
                "target": float(row["target_revenue"]) if row["target_revenue"] is not None else None,
                "realized": float(row["realized_revenue"]) if row["realized_revenue"] is not None else None,
            }

        years = list(range(start_year, end_year + 1))
        year_entries: list[dict[str, Any]] = []
        for year in years:
            meta_series: list[Optional[float]] = []
            real_series: list[Optional[float]] = []
            order_counts: list[Optional[int]] = []
            monthly_diff_pct: list[Optional[float]] = []

            for month in range(1, 13):
                key = (year, month)
                target_month = target_data.get(key) or {}
                meta_value = target_month.get("target")
                order_data = month_data.get(key)
                spreadsheet_realized = target_month.get("realized")

                meta_series.append(meta_value if meta_value is not None else None)
                if order_data:
                    real_value: Optional[float] = float(order_data["revenue"])
                    order_counts.append(int(order_data["orders"]))
                else:
                    real_value = spreadsheet_realized
                    order_counts.append(None)

                real_series.append(real_value if real_value is not None else None)

                if meta_value is None or meta_value == 0 or real_value is None:
                    monthly_diff_pct.append(None)
                else:
                    monthly_diff_pct.append(((real_value - meta_value) / meta_value) * 100)

            meta_total = sum(value for value in meta_series if value is not None) if any(
                value is not None for value in meta_series
            ) else None
            real_total = sum(value for value in real_series if value is not None) if any(
                value is not None for value in real_series
            ) else None
            order_total = sum(value for value in order_counts if value is not None) if any(
                value is not None for value in order_counts
            ) else None
            if meta_total is None or meta_total == 0 or real_total is None:
                diff_pct = None
            else:
                diff_pct = ((real_total - meta_total) / meta_total) * 100

            year_entries.append(
                {
                    "year": year,
                    "meta": meta_series,
                    "realizado": real_series,
                    "orderCounts": order_counts,
                    "orderTotal": order_total,
                    "metaTotal": meta_total,
                    "realTotal": real_total,
                    "diffPct": diff_pct,
                    "monthlyDiffPct": monthly_diff_pct,
                    "metaMissing": [month for month, value in zip(MONTHS, meta_series) if value is None],
                    "realMissing": [month for month, value in zip(MONTHS, real_series) if value is None],
                }
            )

        daily_totals = _daily_totals_for_source(source, start, end)
        monthly_targets = {
            key: float(values["target"])
            for key, values in target_data.items()
            if values.get("target") is not None
        }
        file_updated = DB_PATH.stat().st_mtime if DB_PATH.exists() else datetime.now(timezone.utc).timestamp()
        sync_info = get_sync_status(source)
    finally:
        conn.close()

    return {
        "months": MONTHS,
        "years": year_entries,
        "weekly": _build_weekly_entries(years, daily_totals, monthly_targets),
        "fileUpdatedAt": file_updated,
        "apiYear": date.today().year,
        "apiSource": source,
        "apiUsed": False,
        "apiError": None,
        "localData": True,
        "dataModel": "tiny_api_with_targets",
        "sync": sync_info,
    }

def build_products_payload(
    source: str,
    period: str,
    metric: str,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
) -> dict[str, Any]:
    if metric not in PRODUCT_METRICS:
        raise ValueError("Metric invalida.")
    source_filter_sql, source_params = _source_scope_sql(source)
    orders_source_filter_sql, orders_source_params = _source_scope_sql(source, "orders")

    start, end, period_key, period_label, is_custom = _resolve_period_dates(
        period,
        PRODUCT_PERIODS,
        start_date,
        end_date,
        datetime.now(),
    )

    conn = _db_connect()
    try:
        mix_rows = conn.execute(
            "SELECT payment_method, shipping_method_label, total "
            "FROM orders "
            f"WHERE {source_filter_sql} AND date(order_date) BETWEEN ? AND ? AND {_orders_analytics_filter_sql()}",
            source_params + (start.isoformat(), end.isoformat()),
        ).fetchall()
        total_orders_row = conn.execute(
            "SELECT COUNT(DISTINCT id) AS count FROM orders "
            f"WHERE {source_filter_sql} AND date(order_date) BETWEEN ? AND ? AND {_orders_analytics_filter_sql()}",
            source_params + (start.isoformat(), end.isoformat()),
        ).fetchone()
        total_orders = int(total_orders_row["count"] or 0)

        total_items_row = conn.execute(
            "SELECT COUNT(*) AS count FROM order_items "
            "JOIN orders ON orders.id = order_items.order_id "
            f"WHERE {orders_source_filter_sql} AND date(orders.order_date) BETWEEN ? AND ? AND {_orders_analytics_filter_sql('orders')}",
            orders_source_params + (start.isoformat(), end.isoformat()),
        ).fetchone()
        total_items = int(total_items_row["count"] or 0)

        rows = conn.execute(
            "SELECT "
            "order_items.sku AS sku, "
            "order_items.product_name AS name, "
            "SUM(COALESCE(order_items.quantity, 0)) AS quantity, "
            "SUM(COALESCE(order_items.revenue, 0)) AS revenue "
            "FROM order_items "
            "JOIN orders ON orders.id = order_items.order_id "
            f"WHERE {orders_source_filter_sql} AND date(orders.order_date) BETWEEN ? AND ? AND {_orders_analytics_filter_sql('orders')} "
            "GROUP BY order_items.sku, order_items.product_name",
            orders_source_params + (start.isoformat(), end.isoformat()),
        ).fetchall()
    finally:
        conn.close()

    items = [
        {
            "name": row["name"],
            "sku": row["sku"],
            "quantity": float(row["quantity"] or 0.0),
            "revenue": float(row["revenue"] or 0.0),
        }
        for row in rows
    ]

    items_by_revenue  = sorted(items, key=lambda item: item["revenue"],  reverse=True)[:15]
    items_by_quantity = sorted(items, key=lambda item: item["quantity"], reverse=True)[:15]
    # Keep items sorted by the requested metric for backward compat
    if metric == "revenue":
        items = items_by_revenue
    else:
        items = items_by_quantity

    payment_counter: Counter = Counter()
    payment_revenue: dict[str, float] = {}
    shipping_counter: Counter = Counter()
    for row in mix_rows:
        payment_method = _normalize_payment_label(row["payment_method"])
        if payment_method:
            payment_counter[payment_method] += 1
            payment_revenue[payment_method] = payment_revenue.get(payment_method, 0.0) + float(row["total"] or 0.0)

        shipping_method = str(row["shipping_method_label"] or "").strip()
        if shipping_method:
            shipping_counter[shipping_method] += 1

    payment_total = sum(payment_counter.values())
    payment_mix = [
        {
            "label": label,
            "orders": int(count),
            "sharePct": round((count / payment_total * 100) if payment_total else 0.0, 1),
            "avgTicket": round(payment_revenue.get(label, 0.0) / count, 2) if count else 0.0,
        }
        for label, count in payment_counter.most_common(7)
    ]
    shipping_total = sum(shipping_counter.values())
    shipping_mix = [
        {
            "label": label,
            "orders": int(count),
            "sharePct": round((count / shipping_total * 100) if shipping_total else 0.0, 1),
        }
        for label, count in shipping_counter.most_common(7)
    ]

    return {
        "source": source,
        "period": period_key,
        "periodLabel": period_label,
        "periodIsCustom": is_custom,
        "metric": metric,
        "items": items[:50],
        "itemsByRevenue": items_by_revenue,
        "itemsByQuantity": items_by_quantity,
        "totalOrders": total_orders,
        "totalItems": total_items,
        "paymentMix": payment_mix,
        "shippingMix": shipping_mix,
        "periodStart": start.isoformat(),
        "periodEnd": end.isoformat(),
    }


def build_customers_payload(
    source: str,
    period: str,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
) -> dict[str, Any]:
    source_filter_sql, source_params = _source_scope_sql(source)
    start, end, period_key, period_label, is_custom = _resolve_period_dates(
        period,
        CUSTOMER_PERIODS,
        start_date,
        end_date,
        datetime.now(),
    )

    conn = _db_connect()
    try:
        rows = conn.execute(
            "SELECT "
            "id, order_date, total, customer_code, customer_name, customer_email, customer_document, customer_city, customer_state "
            "FROM orders "
            f"WHERE {source_filter_sql} AND date(order_date) BETWEEN ? AND ? AND {_orders_analytics_filter_sql()}",
            source_params + (start.isoformat(), end.isoformat()),
        ).fetchall()
    finally:
        conn.close()

    customers: dict[str, dict[str, Any]] = {}
    total_revenue = 0.0

    for row in rows:
        total = float(row["total"] or 0.0)
        total_revenue += total

        key = (
            row["customer_email"]
            or row["customer_document"]
            or row["customer_code"]
            or row["customer_name"]
            or f"order:{row['id']}"
        )
        key = str(key).strip()

        entry = customers.get(key)
        if not entry:
            entry = {
                "name": row["customer_name"] or "Cliente",
                "email": row["customer_email"],
                "document": row["customer_document"],
                "city": row["customer_city"],
                "state": row["customer_state"],
                "orders": 0,
                "revenue": 0.0,
                "last_order": None,
            }
            customers[key] = entry

        entry["orders"] += 1
        entry["revenue"] += total

        order_date = _parse_date(row["order_date"])
        if order_date and (entry["last_order"] is None or order_date > entry["last_order"]):
            entry["last_order"] = order_date

    unique_customers = len(customers)
    order_count = len(rows)
    avg_ticket = total_revenue / order_count if order_count else None
    avg_orders_per_customer = order_count / unique_customers if unique_customers else None
    first_time = sum(1 for entry in customers.values() if entry["orders"] == 1)
    repeat_customers = sum(1 for entry in customers.values() if entry["orders"] > 1)

    state_counter: Counter = Counter()
    city_counter: Counter = Counter()
    loyalty_counter: Counter = Counter()
    state_revenue_totals: dict[str, float] = {}
    state_order_totals: Counter = Counter()

    for row in rows:
        state_label = (row["customer_state"] or "").strip().upper()
        if not state_label:
            continue
        total = float(row["total"] or 0.0)
        state_revenue_totals[state_label] = state_revenue_totals.get(state_label, 0.0) + total
        state_order_totals[state_label] += 1

    for entry in customers.values():
        state = (entry.get("state") or "").strip()
        if state:
            state_counter[state] += 1

        city = (entry.get("city") or "").strip()
        city_key = " - ".join(part for part in [city, state] if part)
        if city_key:
            city_counter[city_key] += 1

        loyalty_bucket = _bucket_value(entry["orders"], LOYALTY_BUCKETS)
        if loyalty_bucket:
            loyalty_counter[loyalty_bucket] += 1

    age_distribution = [{"label": label, "count": 0} for label, _, _ in AGE_BUCKETS]
    age_distribution.append({"label": "Sem dados", "count": unique_customers})

    gender_distribution = [
        {"label": "Feminino", "count": 0},
        {"label": "Masculino", "count": 0},
        {"label": "Outro", "count": 0},
        {"label": "Sem dados", "count": unique_customers},
    ]

    loyalty_distribution = [
        {"label": label, "count": loyalty_counter.get(label, 0)} for label, _, _ in LOYALTY_BUCKETS
    ]

    state_revenue_rows = sorted(
        state_revenue_totals.items(),
        key=lambda item: item[1],
        reverse=True,
    )
    state_revenue_total = sum(value for _, value in state_revenue_rows)
    top_limit = 7
    states_revenue: list[dict[str, Any]] = []
    for label, revenue in state_revenue_rows[:top_limit]:
        share = (revenue / state_revenue_total * 100) if state_revenue_total else 0.0
        states_revenue.append(
            {
                "label": label,
                "revenue": round(revenue, 2),
                "orders": int(state_order_totals.get(label, 0)),
                "share": round(share, 1),
            }
        )
    if len(state_revenue_rows) > top_limit:
        other_rows = state_revenue_rows[top_limit:]
        other_revenue = sum(value for _, value in other_rows)
        other_orders = sum(int(state_order_totals.get(label, 0)) for label, _ in other_rows)
        other_share = (other_revenue / state_revenue_total * 100) if state_revenue_total else 0.0
        states_revenue.append(
            {
                "label": "Outros",
                "revenue": round(other_revenue, 2),
                "orders": int(other_orders),
                "share": round(other_share, 1),
            }
        )

    top_customers = sorted(
        [
            {
                "name": entry["name"],
                "email": entry["email"],
                "city": entry["city"],
                "state": entry["state"],
                "orders": entry["orders"],
                "revenue": entry["revenue"],
                "lastOrder": entry["last_order"].isoformat() if entry["last_order"] else None,
            }
            for entry in customers.values()
        ],
        key=lambda item: item["revenue"],
        reverse=True,
    )[:15]

    top_revenue = sum(item["revenue"] for item in top_customers)
    top_share = (top_revenue / total_revenue * 100) if total_revenue else None

    return {
        "source": source,
        "period": period_key,
        "periodLabel": period_label,
        "periodIsCustom": is_custom,
        "periodStart": start.isoformat(),
        "periodEnd": end.isoformat(),
        "orderCount": order_count,
        "revenue": total_revenue,
        "avgTicket": avg_ticket,
        "uniqueCustomers": unique_customers,
        "firstTimeCustomers": first_time,
        "repeatCustomers": repeat_customers,
        "demographics": {
            "age": age_distribution,
            "gender": gender_distribution,
            "states": _distribution_from_counter(state_counter, top=7, include_other=True),
            "statesRevenue": states_revenue,
            "cities": _distribution_from_counter(city_counter, top=7, include_other=True),
        },
        "loyalty": loyalty_distribution,
        "topCustomers": top_customers,
        "topShare": top_share,
        "avgOrdersPerCustomer": avg_orders_per_customer,
    }


def _ga4_property_for_source(source: str) -> Optional[str]:
    if source in {"brew", "brewnh", "brewpoa", "bigb"}:
        return GA4_PROPERTY_ID_BREW
    if source in {"grow", "grow_fisica"}:
        return GA4_PROPERTY_ID_GROW
    return None


def _resolve_credentials_path(raw_value: Optional[str]) -> Optional[Path]:
    if not raw_value:
        return None

    cleaned = str(raw_value).strip().strip('"').strip("'")
    if not cleaned:
        return None

    candidate = Path(cleaned)
    if candidate.exists():
        return candidate

    relative_candidate = (TINY_DIR / cleaned).resolve()
    if relative_candidate.exists():
        return relative_candidate

    normalized = cleaned.replace("\\", "/")

    if normalized.startswith("/mnt/") and len(normalized) > 6 and normalized[5].isalpha() and normalized[6] == "/":
        drive = normalized[5].upper()
        win_path = Path(f"{drive}:{normalized[6:]}")
        if win_path.exists():
            return win_path

    if len(normalized) > 2 and normalized[1] == ":" and normalized[0].isalpha():
        drive = normalized[0].lower()
        rest = normalized[2:]
        unix_like = Path(f"/mnt/{drive}{rest}")
        if unix_like.exists():
            return unix_like

    return candidate


def _ga4_error_message(payload: Any) -> str:
    if not isinstance(payload, dict):
        return "Erro na API GA4."

    error = payload.get("error")
    if isinstance(error, dict):
        msg = error.get("message")
        if msg:
            return str(msg)
    if isinstance(error, str):
        return error
    return "Erro na API GA4."


def _ga4_authorized_session() -> tuple[Optional[Any], Optional[str]]:
    try:
        from google.auth.transport.requests import AuthorizedSession
        from google.oauth2 import service_account
    except Exception:
        return None, "Dependencia ausente: instale google-auth."

    # Prefer inline JSON (base64) — works in cloud without file mounts
    creds_b64 = os.getenv("GOOGLE_CREDENTIALS_B64")
    if creds_b64:
        try:
            import base64, json as _json
            info = _json.loads(base64.b64decode(creds_b64).decode())
            credentials = service_account.Credentials.from_service_account_info(info, scopes=GA4_SCOPES)
            return AuthorizedSession(credentials), None
        except Exception as exc:
            return None, f"Falha ao carregar credenciais GA4 (base64): {exc}"

    # Fallback: file path (local dev)
    credentials_path = _resolve_credentials_path(GOOGLE_APPLICATION_CREDENTIALS)
    if not credentials_path:
        return None, "GOOGLE_APPLICATION_CREDENTIALS nao configurado."
    if not credentials_path.exists():
        return None, f"Arquivo de credenciais nao encontrado: {credentials_path}"

    try:
        credentials = service_account.Credentials.from_service_account_file(
            str(credentials_path),
            scopes=GA4_SCOPES,
        )
        return AuthorizedSession(credentials), None
    except Exception as exc:
        return None, f"Falha ao carregar credenciais GA4: {exc}"


def _ga4_run_report(
    session: Any,
    property_id: str,
    body: dict[str, Any],
) -> tuple[Optional[dict[str, Any]], Optional[str]]:
    url = f"https://analyticsdata.googleapis.com/v1beta/properties/{property_id}:runReport"
    try:
        response = session.post(url, json=body, timeout=REQUEST_TIMEOUT)
    except requests.RequestException as exc:
        return None, f"Falha na API GA4: {exc}"
    except Exception as exc:  # noqa: BLE001
        return None, f"Falha na API GA4: {exc}"

    try:
        payload = response.json()
    except ValueError:
        return None, "Resposta invalida da API GA4."

    if response.status_code >= 400:
        return None, _ga4_error_message(payload)

    return payload, None


def _ga4_metric_value(row: dict[str, Any], index: int = 0) -> float:
    try:
        return float(((row.get("metricValues") or [])[index] or {}).get("value") or 0.0)
    except Exception:
        return 0.0


def _ga4_dimension_value(row: dict[str, Any], index: int = 0) -> str:
    try:
        return str(((row.get("dimensionValues") or [])[index] or {}).get("value") or "").strip()
    except Exception:
        return ""


def _orders_summary_for_period(source: str, start: date, end: date) -> tuple[int, float]:
    conn = _db_connect()
    try:
        row = conn.execute(
            "SELECT COUNT(*) AS orders, SUM(total) AS revenue "
            f"FROM orders WHERE store_code = ? AND date(order_date) BETWEEN ? AND ? AND {_orders_analytics_filter_sql()}",
            (source, start.isoformat(), end.isoformat()),
        ).fetchone()
    finally:
        conn.close()

    return int(row["orders"] or 0), float(row["revenue"] or 0.0)


def _build_funnel_steps(values: dict[str, int]) -> list[dict[str, Any]]:
    steps: list[dict[str, Any]] = []
    first_value = max(int(values.get(FUNNEL_STEP_DEFINITIONS[0]["key"], 0)), 1)
    previous_value: Optional[int] = None

    for definition in FUNNEL_STEP_DEFINITIONS:
        key = definition["key"]
        current = int(values.get(key, 0))
        if previous_value is None:
            stage_rate = 100.0
            dropoff_count = 0
            dropoff_rate = 0.0
        else:
            stage_rate = (current / previous_value * 100) if previous_value > 0 else 0.0
            dropoff_count = max(previous_value - current, 0)
            dropoff_rate = ((dropoff_count / previous_value) * 100) if previous_value > 0 else 0.0

        cumulative_rate = (current / first_value) * 100
        steps.append(
            {
                "key": key,
                "label": definition["label"],
                "description": definition["description"],
                "focus": definition["focus"],
                "value": current,
                "stageRatePct": round(stage_rate, 1),
                "cumulativeRatePct": round(cumulative_rate, 1),
                "dropoffCount": int(dropoff_count),
                "dropoffRatePct": round(dropoff_rate, 1),
            }
        )
        previous_value = current

    return steps


def _sample_funnel_payload(
    source: str,
    period: str,
    start: date,
    end: date,
    warning: str,
) -> dict[str, Any]:
    is_custom = period == "custom"
    period_label = FUNNEL_PERIODS.get(period) or f"{start.isoformat()} a {end.isoformat()}"
    step_values = {
        "sessions": 0,
        "engaged_sessions": 0,
        "add_to_cart": 0,
        "purchase": 0,
    }
    return {
        "source": source,
        "period": period,
        "periodLabel": period_label,
        "periodIsCustom": is_custom,
        "periodStart": start.isoformat(),
        "periodEnd": end.isoformat(),
        "apiUsed": False,
        "isSample": True,
        "warning": warning,
        "steps": _build_funnel_steps(step_values),
        "conversion": 0.0,
        "engagement": 0.0,
        "cartRate": 0.0,
        "ticket": 0.0,
        "topProducts": [],
        "regions": [],
        "campaigns": [],
    }


def _funnel_from_db(
    source: str,
    start: date,
    end: date,
    period_key: str,
    period_label: str,
    is_custom: bool,
) -> Optional[dict[str, Any]]:
    """Aggregate funnel metrics from ga4_daily_metrics. Returns None if no data."""
    conn = _db_connect()
    try:
        row = conn.execute(
            """
            SELECT
                SUM(sessions)         AS sessions,
                SUM(engaged_sessions) AS engaged_sessions,
                SUM(add_to_cart)      AS add_to_cart,
                SUM(purchases)        AS purchases,
                SUM(revenue)          AS revenue,
                COUNT(*)              AS row_count
            FROM ga4_daily_metrics
            WHERE source = ? AND date BETWEEN ? AND ?
            """,
            (source, start.isoformat(), end.isoformat()),
        ).fetchone()
    finally:
        conn.close()

    if not row or (row["row_count"] or 0) == 0:
        return None

    sessions         = int(row["sessions"] or 0)
    engaged_sessions = int(row["engaged_sessions"] or 0)
    add_to_cart      = int(row["add_to_cart"] or 0)
    purchase_events  = int(row["purchases"] or 0)

    orders_count, orders_revenue = _orders_summary_for_period(source, start, end)
    ticket = (orders_revenue / orders_count) if orders_count else None

    conversion = (purchase_events / sessions) if sessions else 0.0
    engagement = (engaged_sessions / sessions) if sessions else 0.0
    cart_rate  = (add_to_cart / sessions) if sessions else 0.0

    step_values = {
        "sessions": sessions,
        "engaged_sessions": engaged_sessions,
        "add_to_cart": add_to_cart,
        "purchase": purchase_events,
    }

    return {
        "source": source,
        "period": period_key,
        "periodLabel": period_label,
        "periodIsCustom": is_custom,
        "periodStart": start.isoformat(),
        "periodEnd": end.isoformat(),
        "apiUsed": False,
        "isSample": False,
        "warning": None,
        "steps": _build_funnel_steps(step_values),
        "conversion": conversion,
        "engagement": engagement,
        "cartRate": cart_rate,
        "ticket": ticket or 0.0,
        "topProducts": [],
        "regions": [],
        "campaigns": [],
    }


def build_customer_retention_payload(
    source: str,
    start_date: str,
    end_date: str,
) -> dict[str, Any]:
    if source not in STORE_RULES and source != "all":
        raise ValueError("Fonte desconhecida.")
    source_filter_sql, source_params = _source_scope_sql(source)
    analytics_filter = _orders_analytics_filter_sql()

    conn = _db_connect()
    try:
        rows = conn.execute(
            f"SELECT customer_email, COUNT(*) AS order_count, SUM(total) AS revenue "
            f"FROM orders "
            f"WHERE {source_filter_sql} "
            f"  AND date(order_date) BETWEEN ? AND ? "
            f"  AND {analytics_filter} "
            f"  AND customer_email IS NOT NULL AND customer_email != '' "
            f"GROUP BY customer_email",
            source_params + (start_date, end_date),
        ).fetchall()
    finally:
        conn.close()

    total_customers = len(rows)
    recurring = [r for r in rows if r["order_count"] > 1]
    new_only = [r for r in rows if r["order_count"] == 1]

    recurring_count = len(recurring)
    new_count = len(new_only)
    recurring_revenue = sum(float(r["revenue"] or 0) for r in recurring)
    new_revenue = sum(float(r["revenue"] or 0) for r in new_only)

    reorder_rate = round(recurring_count / total_customers * 100, 1) if total_customers > 0 else 0.0
    avg_ticket_recurring = round(recurring_revenue / recurring_count, 2) if recurring_count > 0 else 0.0
    avg_ticket_new = round(new_revenue / new_count, 2) if new_count > 0 else 0.0

    return {
        "source": source,
        "periodStart": start_date,
        "periodEnd": end_date,
        "totalCustomers": total_customers,
        "recurringCustomers": recurring_count,
        "newCustomers": new_count,
        "reorderRatePct": reorder_rate,
        "avgTicketRecurring": avg_ticket_recurring,
        "avgTicketNew": avg_ticket_new,
    }


def build_customer_churn_payload(
    source: str,
    inactive_days: int = 30,
) -> dict[str, Any]:
    if source not in STORE_RULES and source != "all":
        raise ValueError("Fonte desconhecida.")
    source_filter_sql, source_params = _source_scope_sql(source)
    analytics_filter = _orders_analytics_filter_sql()
    cutoff_date = (date.today() - timedelta(days=inactive_days)).isoformat()

    conn = _db_connect()
    try:
        rows = conn.execute(
            f"SELECT customer_email, customer_name, "
            f"  MAX(order_date) AS last_order_date, "
            f"  COUNT(*) AS total_orders, "
            f"  SUM(total) AS total_revenue "
            f"FROM orders "
            f"WHERE {source_filter_sql} "
            f"  AND {analytics_filter} "
            f"  AND customer_email IS NOT NULL AND customer_email != '' "
            f"GROUP BY customer_email "
            f"HAVING MAX(order_date) < ? "
            f"ORDER BY total_revenue DESC",
            source_params + (cutoff_date,),
        ).fetchall()
    finally:
        conn.close()

    churned_count = len(rows)
    top_churned = [
        {
            "customerEmail": row["customer_email"],
            "customerName": row["customer_name"],
            "lastOrderDate": row["last_order_date"],
            "totalOrders": int(row["total_orders"]),
            "totalRevenue": float(row["total_revenue"] or 0),
        }
        for row in rows[:20]
    ]

    return {
        "source": source,
        "inactiveDays": inactive_days,
        "cutoffDate": cutoff_date,
        "churnedCount": churned_count,
        "topChurnedCustomers": top_churned,
    }


def sync_ga4_daily(
    target_date: Optional[str] = None,
    sources: Optional[list[str]] = None,
) -> dict[str, Any]:
    """Fetch GA4 metrics for target_date (default: yesterday) and upsert into ga4_daily_metrics.

    Uses the existing _ga4_authorized_session / _ga4_run_report helpers (REST API via google-auth).
    Returns dict: {source: {"status": "ok"|"skip"|"error", "date": ..., "rowsUpserted": int}}
    """
    if target_date is None:
        target_date = (date.today() - timedelta(days=1)).isoformat()

    if sources is None:
        sources = []
        if GA4_PROPERTY_ID_BREW:
            sources.append("brew")
        if GA4_PROPERTY_ID_GROW:
            sources.append("grow")

    if not sources:
        logger.warning("sync_ga4_daily: nenhum GA4_PROPERTY_ID configurado, pulando.")
        return {}

    session, err = _ga4_authorized_session()
    if session is None:
        reason = err or "client_init_failed"
        logger.warning("sync_ga4_daily: sessao GA4 nao disponivel — %s", reason)
        return {s: {"status": "skip", "reason": reason} for s in sources}

    property_map: dict[str, Optional[str]] = {
        "brew": GA4_PROPERTY_ID_BREW,
        "grow": GA4_PROPERTY_ID_GROW,
    }

    results: dict[str, Any] = {}

    for source in sources:
        prop_id = property_map.get(source)
        if not prop_id:
            results[source] = {"status": "skip", "reason": "no_property_id"}
            continue

        try:
            body = {
                "dimensions": [{"name": "date"}],
                "metrics": [
                    {"name": "sessions"},
                    {"name": "totalUsers"},
                    {"name": "newUsers"},
                    {"name": "engagedSessions"},
                    {"name": "addToCarts"},
                    {"name": "checkouts"},
                    {"name": "purchases"},
                    {"name": "purchaseRevenue"},
                ],
                "dateRanges": [{"startDate": target_date, "endDate": target_date}],
            }

            payload, api_err = _ga4_run_report(session, prop_id, body)
            if api_err:
                logger.error("sync_ga4_daily: %s %s — %s", source, target_date, api_err)
                results[source] = {"status": "error", "date": target_date, "error": api_err}
                continue

            rows_data = (payload or {}).get("rows") or []
            rows_upserted = 0
            conn = _db_connect()
            try:
                for row in rows_data:
                    date_val = _ga4_dimension_value(row, 0)
                    # GA4 returns date as YYYYMMDD — normalise to YYYY-MM-DD
                    if len(date_val) == 8 and date_val.isdigit():
                        date_val = f"{date_val[:4]}-{date_val[4:6]}-{date_val[6:]}"

                    def _int(idx: int) -> int:
                        return int(_ga4_metric_value(row, idx))

                    def _float(idx: int) -> float:
                        return _ga4_metric_value(row, idx)

                    conn.execute(
                        """
                        INSERT INTO ga4_daily_metrics
                            (source, date, sessions, users, new_users, engaged_sessions,
                             add_to_cart, checkouts, purchases, revenue)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        ON CONFLICT(source, date) DO UPDATE SET
                            sessions         = excluded.sessions,
                            users            = excluded.users,
                            new_users        = excluded.new_users,
                            engaged_sessions = excluded.engaged_sessions,
                            add_to_cart      = excluded.add_to_cart,
                            checkouts        = excluded.checkouts,
                            purchases        = excluded.purchases,
                            revenue          = excluded.revenue
                        """,
                        (
                            source,
                            date_val,
                            _int(0),
                            _int(1),
                            _int(2),
                            _int(3),
                            _int(4),
                            _int(5),
                            _int(6),
                            _float(7),
                        ),
                    )
                    rows_upserted += 1
                conn.commit()
            finally:
                conn.close()

            results[source] = {"status": "ok", "date": target_date, "rowsUpserted": rows_upserted}
            logger.info(
                "sync_ga4_daily: %s %s → %d rows upserted",
                source,
                target_date,
                rows_upserted,
            )

        except Exception as exc:  # noqa: BLE001
            logger.error("sync_ga4_daily: %s %s falhou: %s", source, target_date, exc)
            results[source] = {"status": "error", "date": target_date, "error": str(exc)}

    return results


def sync_ga4_range(
    start_date: str,
    end_date: str,
    sources: Optional[list[str]] = None,
) -> dict[str, Any]:
    """Fetch GA4 daily metrics for a date range and upsert into ga4_daily_metrics.

    Makes one API call per source with date dimension — efficient for bulk historical loads.
    Returns {source: {"status": "ok"|"skip"|"error", "dateRange": ..., "rowsUpserted": int}}.
    """
    if sources is None:
        sources = []
        if GA4_PROPERTY_ID_BREW:
            sources.append("brew")
        if GA4_PROPERTY_ID_GROW:
            sources.append("grow")

    if not sources:
        logger.warning("sync_ga4_range: nenhum GA4_PROPERTY_ID configurado.")
        return {}

    session, err = _ga4_authorized_session()
    if session is None:
        reason = err or "client_init_failed"
        logger.warning("sync_ga4_range: sessao GA4 indisponivel — %s", reason)
        return {s: {"status": "skip", "reason": reason} for s in sources}

    property_map: dict[str, Optional[str]] = {
        "brew": GA4_PROPERTY_ID_BREW,
        "grow": GA4_PROPERTY_ID_GROW,
    }

    results: dict[str, Any] = {}

    for source in sources:
        prop_id = property_map.get(source)
        if not prop_id:
            results[source] = {"status": "skip", "reason": "no_property_id"}
            continue

        try:
            body = {
                "dimensions": [{"name": "date"}],
                "metrics": [
                    {"name": "sessions"},
                    {"name": "totalUsers"},
                    {"name": "newUsers"},
                    {"name": "engagedSessions"},
                    {"name": "addToCarts"},
                    {"name": "checkouts"},
                    {"name": "ecommercePurchases"},
                    {"name": "purchaseRevenue"},
                ],
                "dateRanges": [{"startDate": start_date, "endDate": end_date}],
                "limit": "500",
            }

            payload, api_err = _ga4_run_report(session, prop_id, body)
            if api_err:
                logger.error("sync_ga4_range: %s %s→%s — %s", source, start_date, end_date, api_err)
                results[source] = {"status": "error", "dateRange": f"{start_date}/{end_date}", "error": api_err}
                continue

            rows_data = (payload or {}).get("rows") or []
            rows_upserted = 0
            conn = _db_connect()
            try:
                for row in rows_data:
                    date_val = _ga4_dimension_value(row, 0)
                    if len(date_val) == 8 and date_val.isdigit():
                        date_val = f"{date_val[:4]}-{date_val[4:6]}-{date_val[6:]}"

                    def _int(idx: int) -> int:
                        return int(_ga4_metric_value(row, idx))

                    def _float(idx: int) -> float:
                        return _ga4_metric_value(row, idx)

                    conn.execute(
                        """
                        INSERT INTO ga4_daily_metrics
                            (source, date, sessions, users, new_users, engaged_sessions,
                             add_to_cart, checkouts, purchases, revenue)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        ON CONFLICT(source, date) DO UPDATE SET
                            sessions         = excluded.sessions,
                            users            = excluded.users,
                            new_users        = excluded.new_users,
                            engaged_sessions = excluded.engaged_sessions,
                            add_to_cart      = excluded.add_to_cart,
                            checkouts        = excluded.checkouts,
                            purchases        = excluded.purchases,
                            revenue          = excluded.revenue
                        """,
                        (
                            source, date_val,
                            _int(0), _int(1), _int(2), _int(3),
                            _int(4), _int(5), _int(6), _float(7),
                        ),
                    )
                    rows_upserted += 1
                conn.commit()
            finally:
                conn.close()

            results[source] = {
                "status": "ok",
                "dateRange": f"{start_date}/{end_date}",
                "rowsUpserted": rows_upserted,
            }
            logger.info(
                "sync_ga4_range: %s %s→%s → %d rows upserted",
                source, start_date, end_date, rows_upserted,
            )

        except Exception as exc:  # noqa: BLE001
            logger.error("sync_ga4_range: %s falhou: %s", source, exc)
            results[source] = {"status": "error", "dateRange": f"{start_date}/{end_date}", "error": str(exc)}

    return results


def _ga4_channel_breakdown(
    session: Any,
    property_id: str,
    date_range: list[dict[str, str]],
) -> list[dict[str, Any]]:
    """Query sessions + funnel events by sessionDefaultChannelGroup.

    Returns a list ordered by CHANNEL_ORDER, each entry with per-step pct breakdown:
    [{"key", "label", "color", "pctByStep": {"sessions": 0.42, "engaged_sessions": 0.44, ...}}]
    """
    report, err = _ga4_run_report(
        session,
        property_id,
        {
            "dimensions": [{"name": "sessionDefaultChannelGroup"}],
            "metrics": [
                {"name": "sessions"},
                {"name": "engagedSessions"},
                {"name": "addToCarts"},
                {"name": "ecommercePurchases"},
            ],
            "dateRanges": date_range,
            "limit": "50",
        },
    )
    if err or not report:
        return []

    # Accumulate per canonical key
    agg: dict[str, dict[str, int]] = {}
    for row in report.get("rows") or []:
        ga4_label = _ga4_dimension_value(row, 0)
        key, _, _ = CHANNEL_GROUPS.get(ga4_label, ("outros", "", ""))
        entry = agg.setdefault(key, {"sessions": 0, "engaged_sessions": 0, "add_to_cart": 0, "purchase": 0})
        entry["sessions"]         += int(_ga4_metric_value(row, 0))
        entry["engaged_sessions"] += int(_ga4_metric_value(row, 1))
        entry["add_to_cart"]      += int(_ga4_metric_value(row, 2))
        entry["purchase"]         += int(_ga4_metric_value(row, 3))

    if not agg:
        return []

    totals = {k: sum(v[k] for v in agg.values()) for k in ("sessions", "engaged_sessions", "add_to_cart", "purchase")}

    result = []
    for key in CHANNEL_ORDER:
        if key not in agg:
            continue
        vals = agg[key]
        label, color = CHANNEL_META[key]
        result.append({
            "key": key,
            "label": label,
            "color": color,
            "pctByStep": {
                step: (vals[step] / totals[step]) if totals[step] else 0.0
                for step in ("sessions", "engaged_sessions", "add_to_cart", "purchase")
            },
        })

    return result


def build_funnel_payload(
    source: str,
    period: str,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
) -> dict[str, Any]:
    if source not in STORE_RULES:
        raise ValueError("Fonte desconhecida.")
    start, end, period_key, period_label, is_custom = _resolve_period_dates(
        period,
        FUNNEL_PERIODS,
        start_date,
        end_date,
        datetime.now(),
    )

    # Check DB cache for base funnel metrics
    db_payload = _funnel_from_db(source, start, end, period_key, period_label, is_custom)

    property_id = _ga4_property_for_source(source)
    date_range = [{"startDate": start.isoformat(), "endDate": end.isoformat()}]

    if not property_id:
        if db_payload is not None:
            db_payload["channelPcts"] = []
            return db_payload
        return _sample_funnel_payload(
            source, period_key, start, end, "Property GA4 nao configurada para esta fonte.",
        )

    session, session_error = _ga4_authorized_session()
    if session_error or not session:
        if db_payload is not None:
            db_payload["channelPcts"] = []
            return db_payload
        return _sample_funnel_payload(
            source, period_key, start, end, session_error or "Falha na autenticacao GA4.",
        )

    # Channel breakdown always fetched live (fast single query)
    channel_pcts = _ga4_channel_breakdown(session, property_id, date_range)

    if db_payload is not None:
        db_payload["channelPcts"] = channel_pcts
        return db_payload

    summary_report, error = _ga4_run_report(
        session,
        property_id,
        {
            "dateRanges": date_range,
            "metrics": [{"name": "sessions"}, {"name": "engagedSessions"}],
        },
    )
    if error or not summary_report:
        return _sample_funnel_payload(
            source,
            period_key,
            start,
            end,
            f"GA4 indisponivel: {error or 'sem resposta'}",
        )

    event_report, error = _ga4_run_report(
        session,
        property_id,
        {
            "dateRanges": date_range,
            "dimensions": [{"name": "eventName"}],
            "metrics": [{"name": "eventCount"}],
            "dimensionFilter": {
                "filter": {
                    "fieldName": "eventName",
                    "inListFilter": {
                        "values": ["add_to_cart", "purchase"],
                    },
                }
            },
            "limit": "100",
        },
    )
    if error or not event_report:
        return _sample_funnel_payload(
            source,
            period_key,
            start,
            end,
            f"GA4 indisponivel: {error or 'sem resposta'}",
        )

    summary_rows = summary_report.get("rows") or []
    summary_row = summary_rows[0] if summary_rows else {}
    sessions = int(round(_ga4_metric_value(summary_row, 0)))
    engaged_sessions = int(round(_ga4_metric_value(summary_row, 1)))

    event_counts: dict[str, int] = {}
    for row in event_report.get("rows") or []:
        name = _ga4_dimension_value(row, 0)
        if not name:
            continue
        event_counts[name] = int(round(_ga4_metric_value(row, 0)))

    add_to_cart = event_counts.get("add_to_cart", 0)
    purchase_events = event_counts.get("purchase", 0)

    orders_count, orders_revenue = _orders_summary_for_period(source, start, end)
    ticket = (orders_revenue / orders_count) if orders_count else None

    conversion = (purchase_events / sessions) if sessions else 0.0
    engagement = (engaged_sessions / sessions) if sessions else 0.0
    cart_rate = (add_to_cart / sessions) if sessions else 0.0

    warnings: list[str] = []

    top_products: list[dict[str, Any]] = []
    product_report, error = _ga4_run_report(
        session,
        property_id,
        {
            "dateRanges": date_range,
            "dimensions": [{"name": "itemName"}],
            "metrics": [{"name": "itemsViewed"}],
            "orderBys": [{"metric": {"metricName": "itemsViewed"}, "desc": True}],
            "limit": "5",
        },
    )
    if error or not product_report:
        warnings.append(f"Produtos mais vistos: {error or 'sem resposta'}")
    else:
        for row in product_report.get("rows") or []:
            name = _ga4_dimension_value(row, 0)
            if not name or name in {"(not set)", "(not provided)"}:
                continue
            top_products.append(
                {
                    "label": name,
                    "count": int(round(_ga4_metric_value(row, 0))),
                }
            )

    regions: list[dict[str, Any]] = []
    region_report, error = _ga4_run_report(
        session,
        property_id,
        {
            "dateRanges": date_range,
            "dimensions": [{"name": "region"}],
            "metrics": [{"name": "sessions"}],
            "orderBys": [{"metric": {"metricName": "sessions"}, "desc": True}],
            "limit": "5",
        },
    )
    if error or not region_report:
        warnings.append(f"Regioes: {error or 'sem resposta'}")
    else:
        total_region_sessions = 0
        tmp_regions: list[tuple[str, int]] = []
        for row in region_report.get("rows") or []:
            label = _ga4_dimension_value(row, 0)
            if not label or label in {"(not set)", "(not provided)"}:
                continue
            count = int(round(_ga4_metric_value(row, 0)))
            total_region_sessions += count
            tmp_regions.append((label, count))

        for label, count in tmp_regions:
            share = (count / total_region_sessions * 100) if total_region_sessions else 0.0
            regions.append({"label": label, "share": round(share, 1)})

    campaigns: list[dict[str, Any]] = []
    campaign_sessions_report, error_sessions = _ga4_run_report(
        session,
        property_id,
        {
            "dateRanges": date_range,
            "dimensions": [{"name": "sessionSourceMedium"}],
            "metrics": [{"name": "sessions"}],
            "orderBys": [{"metric": {"metricName": "sessions"}, "desc": True}],
            "limit": "8",
        },
    )
    campaign_purchase_report, error_purchase = _ga4_run_report(
        session,
        property_id,
        {
            "dateRanges": date_range,
            "dimensions": [{"name": "sessionSourceMedium"}],
            "metrics": [{"name": "eventCount"}],
            "dimensionFilter": {
                "filter": {
                    "fieldName": "eventName",
                    "stringFilter": {"matchType": "EXACT", "value": "purchase"},
                }
            },
            "limit": "50",
        },
    )
    if error_sessions or not campaign_sessions_report:
        warnings.append(f"Campanhas (sessoes): {error_sessions or 'sem resposta'}")
    elif error_purchase or not campaign_purchase_report:
        warnings.append(f"Campanhas (conversao): {error_purchase or 'sem resposta'}")
    else:
        purchases_by_source: dict[str, int] = {}
        for row in campaign_purchase_report.get("rows") or []:
            label = _ga4_dimension_value(row, 0)
            if not label:
                continue
            purchases_by_source[label] = int(round(_ga4_metric_value(row, 0)))

        for row in campaign_sessions_report.get("rows") or []:
            label = _ga4_dimension_value(row, 0)
            if not label or label in {"(not set)", "(not provided)"}:
                continue
            sessions_count = int(round(_ga4_metric_value(row, 0)))
            purchase_count = purchases_by_source.get(label, 0)
            conv_pct = (purchase_count / sessions_count * 100) if sessions_count else 0.0
            source_label, medium_label = (part.strip() for part in (label.split("/", 1) + [""])[:2])
            campaigns.append(
                {
                    "label": label,
                    "source": source_label,
                    "medium": medium_label,
                    "conv": round(conv_pct, 1),
                    "sessions": sessions_count,
                    "purchases": purchase_count,
                }
            )

    warning = " | ".join(warnings) if warnings else None

    step_values = {
        "sessions": sessions,
        "engaged_sessions": engaged_sessions,
        "add_to_cart": add_to_cart,
        "purchase": purchase_events,
    }

    return {
        "source": source,
        "period": period_key,
        "periodLabel": period_label,
        "periodIsCustom": is_custom,
        "periodStart": start.isoformat(),
        "periodEnd": end.isoformat(),
        "apiUsed": True,
        "isSample": False,
        "warning": warning,
        "steps": _build_funnel_steps(step_values),
        "conversion": conversion,
        "engagement": engagement,
        "cartRate": cart_rate,
        "ticket": ticket or 0.0,
        "topProducts": top_products,
        "regions": regions,
        "campaigns": campaigns,
        "channelPcts": channel_pcts,
    }


def _calc_delta_pct(current: Optional[float], previous: Optional[float]) -> Optional[float]:
    if current is None or previous is None or previous == 0:
        return None
    return ((current - previous) / previous) * 100


def _metric_payload(value: Optional[float], previous: Optional[float]) -> dict[str, Any]:
    return {
        "value": value,
        "previous": previous,
        "deltaPct": _calc_delta_pct(value, previous),
    }


def _format_currency_brl(value: Optional[float]) -> str:
    if value is None:
        return "-"
    rounded = round(float(value), 2)
    formatted = f"{rounded:,.2f}"
    formatted = formatted.replace(",", "X").replace(".", ",").replace("X", ".")
    return f"R$ {formatted}"


def _format_number(value: Optional[float], digits: int = 1) -> str:
    if value is None:
        return "-"
    pattern = f"{{:.{digits}f}}"
    return pattern.format(float(value)).replace(".", ",")


def _snapshot_period(
    conn: sqlite3.Connection,
    source: str,
    start: date,
    end: date,
) -> dict[str, Any]:
    rows = conn.execute(
        "SELECT "
        "id, total, customer_code, customer_name, customer_email, customer_document, customer_state, payment_method, shipping_method_label "
        "FROM orders "
        f"WHERE store_code = ? AND date(order_date) BETWEEN ? AND ? AND {_orders_analytics_filter_sql()}",
        (source, start.isoformat(), end.isoformat()),
    ).fetchall()

    revenue = 0.0
    customer_totals: dict[str, dict[str, Any]] = {}
    state_revenue_totals: dict[str, float] = {}
    payment_counter: Counter = Counter()
    shipping_counter: Counter = Counter()

    for row in rows:
        total = float(row["total"] or 0.0)
        revenue += total

        customer_key = (
            row["customer_email"]
            or row["customer_document"]
            or row["customer_code"]
            or row["customer_name"]
            or f"order:{row['id']}"
        )
        customer_key = str(customer_key).strip()

        customer_entry = customer_totals.get(customer_key)
        if not customer_entry:
            customer_entry = {"orders": 0, "revenue": 0.0}
            customer_totals[customer_key] = customer_entry
        customer_entry["orders"] += 1
        customer_entry["revenue"] += total

        state = str(row["customer_state"] or "").strip().upper()
        if state:
            state_revenue_totals[state] = state_revenue_totals.get(state, 0.0) + total

        payment_method = _normalize_payment_label(row["payment_method"])
        if payment_method:
            payment_counter[payment_method] += 1

        shipping_method = str(row["shipping_method_label"] or "").strip()
        if shipping_method:
            shipping_counter[shipping_method] += 1

    order_count = len(rows)
    unique_customers = len(customer_totals)
    repeat_customers = sum(1 for item in customer_totals.values() if int(item["orders"]) > 1)
    avg_ticket = (revenue / order_count) if order_count else None
    items_row = conn.execute(
        "SELECT SUM(COALESCE(order_items.quantity, 0)) AS quantity "
        "FROM order_items "
        "JOIN orders ON orders.id = order_items.order_id "
        f"WHERE orders.store_code = ? AND date(orders.order_date) BETWEEN ? AND ? AND {_orders_analytics_filter_sql('orders')}",
        (source, start.isoformat(), end.isoformat()),
    ).fetchone()
    items_sold = float(items_row["quantity"] or 0.0)
    items_per_order = (items_sold / order_count) if order_count else None
    repeat_rate = (repeat_customers / unique_customers * 100) if unique_customers else None
    revenue_per_customer = (revenue / unique_customers) if unique_customers else None

    product_rows = conn.execute(
        "SELECT "
        "order_items.product_name AS name, "
        "order_items.sku AS sku, "
        "SUM(COALESCE(order_items.quantity, 0)) AS quantity, "
        "SUM(COALESCE(order_items.revenue, 0)) AS revenue "
        "FROM order_items "
        "JOIN orders ON orders.id = order_items.order_id "
        f"WHERE orders.store_code = ? AND date(orders.order_date) BETWEEN ? AND ? AND {_orders_analytics_filter_sql('orders')} "
        "GROUP BY order_items.product_name, order_items.sku "
        "ORDER BY revenue DESC, quantity DESC "
        "LIMIT 7",
        (source, start.isoformat(), end.isoformat()),
    ).fetchall()
    top_products = [
        {
            "name": row["name"] or "Produto sem nome",
            "sku": row["sku"],
            "quantity": float(row["quantity"] or 0.0),
            "revenue": float(row["revenue"] or 0.0),
        }
        for row in product_rows
    ]

    state_total = sum(state_revenue_totals.values())
    top_states = []
    for label, state_revenue in sorted(
        state_revenue_totals.items(),
        key=lambda item: item[1],
        reverse=True,
    )[:7]:
        share = (state_revenue / state_total * 100) if state_total else 0.0
        top_states.append(
            {
                "label": label,
                "revenue": round(state_revenue, 2),
                "sharePct": round(share, 1),
            }
        )

    payment_total = sum(payment_counter.values())
    payment_mix = []
    for label, count in payment_counter.most_common(7):
        share = (count / payment_total * 100) if payment_total else 0.0
        payment_mix.append(
            {
                "label": label,
                "orders": int(count),
                "sharePct": round(share, 1),
            }
        )

    shipping_total = sum(shipping_counter.values())
    shipping_mix = []
    for label, count in shipping_counter.most_common(7):
        share = (count / shipping_total * 100) if shipping_total else 0.0
        shipping_mix.append(
            {
                "label": label,
                "orders": int(count),
                "sharePct": round(share, 1),
            }
        )

    return {
        "orderCount": order_count,
        "revenue": revenue,
        "avgTicket": avg_ticket,
        "uniqueCustomers": unique_customers,
        "repeatCustomers": repeat_customers,
        "repeatRatePct": repeat_rate,
        "itemsSold": items_sold,
        "itemsPerOrder": items_per_order,
        "revenuePerCustomer": revenue_per_customer,
        "topProducts": top_products,
        "topStates": top_states,
        "paymentMix": payment_mix,
        "shippingMix": shipping_mix,
    }


def _operational_snapshot(conn: sqlite3.Connection) -> dict[str, Any]:
    max_rows = 50
    counts_row = conn.execute(
        "SELECT "
        "SUM(CASE WHEN issue_missing_link_marker = 1 THEN 1 ELSE 0 END) AS missing_link_marker_count, "
        "SUM(CASE WHEN issue_open_over_2_days = 1 THEN 1 ELSE 0 END) AS open_over_2_days_count "
        "FROM operational_watch_orders "
        "WHERE active_issue = 1",
    ).fetchone()

    missing_rows = conn.execute(
        "SELECT order_number, order_date, customer_name, situacao "
        "FROM operational_watch_orders "
        "WHERE issue_missing_link_marker = 1 AND active_issue = 1 "
        "ORDER BY date(order_date) DESC, tiny_id DESC "
        "LIMIT ?",
        (max_rows,),
    ).fetchall()
    open_rows = conn.execute(
        "SELECT order_number, order_date, customer_name, situacao "
        "FROM operational_watch_orders "
        "WHERE issue_open_over_2_days = 1 AND active_issue = 1 "
        "ORDER BY date(order_date) ASC, tiny_id ASC "
        "LIMIT ?",
        (max_rows,),
    ).fetchall()

    missing_by_store = conn.execute(
        "SELECT o.store_code, COUNT(*) as cnt "
        "FROM operational_watch_orders ow "
        "JOIN orders o ON o.tiny_id = ow.tiny_id "
        "WHERE ow.issue_missing_link_marker = 1 AND ow.active_issue = 1 "
        "GROUP BY o.store_code"
    ).fetchall()

    open_by_store = conn.execute(
        "SELECT o.store_code, COUNT(*) as cnt "
        "FROM operational_watch_orders ow "
        "JOIN orders o ON o.tiny_id = ow.tiny_id "
        "WHERE ow.issue_open_over_2_days = 1 AND ow.active_issue = 1 "
        "GROUP BY o.store_code"
    ).fetchall()

    return {
        "missingLinkMarkerCount": int((counts_row["missing_link_marker_count"] if counts_row else 0) or 0),
        "openOverNDaysCount": int((counts_row["open_over_2_days_count"] if counts_row else 0) or 0),
        "missingLinkMarkerOrders": [
            {
                "orderNumber": str(row["order_number"] or ""),
                "orderDate": row["order_date"],
                "customerName": row["customer_name"],
                "situacao": row["situacao"],
            }
            for row in missing_rows
        ],
        "openOverNDaysOrders": [
            {
                "orderNumber": str(row["order_number"] or ""),
                "orderDate": row["order_date"],
                "customerName": row["customer_name"],
                "situacao": row["situacao"],
            }
            for row in open_rows
        ],
        "missingLinkMarkerByStore": {row["store_code"]: row["cnt"] for row in missing_by_store},
        "openOverNDaysByStore": {row["store_code"]: row["cnt"] for row in open_by_store},
    }


def _build_management_lists(
    view_mode: str,
    current_snapshot: dict[str, Any],
    previous_snapshot: dict[str, Any],
    period_label: str,
    funnel_summary: Optional[dict[str, Any]],
    operational_snapshot: Optional[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    current_revenue = float(current_snapshot.get("revenue") or 0.0)
    previous_revenue = float(previous_snapshot.get("revenue") or 0.0)
    revenue_delta = _calc_delta_pct(current_revenue, previous_revenue)

    current_orders = int(current_snapshot.get("orderCount") or 0)
    previous_orders = int(previous_snapshot.get("orderCount") or 0)
    orders_delta = _calc_delta_pct(float(current_orders), float(previous_orders))

    current_ticket = current_snapshot.get("avgTicket")
    previous_ticket = previous_snapshot.get("avgTicket")
    ticket_delta = _calc_delta_pct(
        float(current_ticket) if current_ticket is not None else None,
        float(previous_ticket) if previous_ticket is not None else None,
    )

    top_state = (current_snapshot.get("topStates") or [{}])[0]
    top_state_label = top_state.get("label") or "-"
    top_state_share = top_state.get("sharePct")

    insights: list[dict[str, Any]] = [
        {
            "title": "Ritmo comercial",
            "detail": (
                f"No período ({period_label}), foram {current_orders} pedidos com "
                f"{_format_currency_brl(current_revenue)} de receita."
            ),
            "value": f"{_format_number(revenue_delta)}%" if revenue_delta is not None else "-",
        },
        {
            "title": "Ticket médio",
            "detail": (
                f"Ticket atual em {_format_currency_brl(current_ticket)} "
                f"vs {_format_currency_brl(previous_ticket)} no período anterior."
            ),
            "value": f"{_format_number(ticket_delta)}%" if ticket_delta is not None else "-",
        },
        {
            "title": "Concentração geográfica",
            "detail": (
                f"UF com maior faturamento: {top_state_label} "
                f"({(_format_number(top_state_share) + '%') if top_state_share is not None else '-'})"
            ),
            "value": "UF líder",
        },
    ]

    alerts: list[dict[str, Any]] = []
    actions: list[dict[str, Any]] = []
    missing_link_marker_count = int((operational_snapshot or {}).get("missingLinkMarkerCount") or 0)
    open_over_2_days_count = int((operational_snapshot or {}).get("openOverNDaysCount") or 0)

    if revenue_delta is not None and revenue_delta < -5:
        alerts.append(
            {
                "title": "Queda relevante de receita",
                "detail": (
                    f"Receita caiu {abs(revenue_delta):.1f}% contra período anterior."
                ),
                "severity": "Alta",
            }
        )
        actions.append(
            {
                "title": "Ativar plano de recuperação de receita",
                "detail": "Reforçar oferta dos top SKUs e revisar campanha/roteiro comercial em até 48h.",
                "priority": "Alta",
            }
        )

    if orders_delta is not None and orders_delta < -5:
        alerts.append(
            {
                "title": "Volume de pedidos em retração",
                "detail": f"Pedidos recuaram {abs(orders_delta):.1f}% no comparativo.",
                "severity": "Média",
            }
        )

    if missing_link_marker_count > 0:
        alerts.append(
            {
                "title": "Pedidos sem vínculo e sem marcador",
                "detail": (
                    f"Foram detectados {missing_link_marker_count} pedido(s) sem vínculo de e-commerce "
                    "e sem uma das tags obrigatórias (BREWNH, BREWPOA, GROW, BIGB)."
                ),
                "severity": "Alta",
            }
        )
        actions.append(
            {
                "title": "Corrigir marcação operacional dos pedidos",
                "detail": "Ajustar pedidos sem vínculo para conter marcador obrigatório e evitar distorção de loja.",
                "priority": "Alta",
            }
        )

    if open_over_2_days_count > 0:
        alerts.append(
            {
                "title": f"Pedidos em aberto há mais de {OPEN_ORDER_ALERT_DAYS} dia(s)",
                "detail": f"Existem {open_over_2_days_count} pedido(s) com status 'Em aberto' fora do SLA.",
                "severity": "Alta",
            }
        )
        actions.append(
            {
                "title": "Atuar no backlog de pedidos em aberto",
                "detail": f"Priorizar conferência e fechamento dos pedidos em aberto acima de {OPEN_ORDER_ALERT_DAYS} dia(s).",
                "priority": "Alta",
            }
        )

    if view_mode == "ecommerce":
        conversion_pct = None
        engagement_pct = None
        if funnel_summary and funnel_summary.get("available"):
            conversion_pct = funnel_summary.get("conversionRatePct")
            engagement_pct = funnel_summary.get("engagementRatePct")
            insights.append(
                {
                    "title": "Eficiência de conversão",
                    "detail": (
                        f"Conversão final em {_format_number(conversion_pct)}% e "
                        f"engajamento em {_format_number(engagement_pct)}%."
                    ),
                    "value": f"{_format_number(conversion_pct)}%" if conversion_pct is not None else "-",
                }
            )

        if funnel_summary and funnel_summary.get("isSample"):
            alerts.append(
                {
                    "title": "Funil em modo de amostra",
                    "detail": "GA4 não retornou dados reais para o período; o dashboard usou valores de fallback.",
                    "severity": "Média",
                }
            )
            actions.append(
                {
                    "title": "Validar conector GA4",
                    "detail": "Revisar credenciais, permissões e propriedade para publicar dados reais de tráfego.",
                    "priority": "Alta",
                }
            )
        elif conversion_pct is not None and conversion_pct < 1.2:
            alerts.append(
                {
                    "title": "Conversão abaixo do esperado",
                    "detail": f"Conversão final em {_format_number(conversion_pct)}%, abaixo do piso sugerido (1,2%).",
                    "severity": "Alta",
                }
            )
            actions.append(
                {
                    "title": "Reduzir abandono de carrinho",
                    "detail": "Priorizar recuperação de checkout e incentivos de frete no funil final.",
                    "priority": "Alta",
                }
            )
        else:
            actions.append(
                {
                    "title": "Escalar canais de melhor retorno",
                    "detail": "Realocar verba para campanhas com maior conversão e maior sessão engajada.",
                    "priority": "Média",
                }
            )
    else:
        repeat_rate = current_snapshot.get("repeatRatePct")
        items_per_order = current_snapshot.get("itemsPerOrder")
        insights.append(
            {
                "title": "Recorrência de clientes",
                "detail": (
                    f"Taxa de clientes recorrentes: {_format_number(repeat_rate)}% "
                    f"no período monitorado."
                ),
                "value": f"{_format_number(repeat_rate)}%" if repeat_rate is not None else "-",
            }
        )

        if repeat_rate is not None and repeat_rate < 22:
            alerts.append(
                {
                    "title": "Baixa recorrência em loja física",
                    "detail": f"Recorrência em {_format_number(repeat_rate)}%, indicando baixa frequência de recompra.",
                    "severity": "Média",
                }
            )
            actions.append(
                {
                    "title": "Programa de retorno de clientes",
                    "detail": "Criar ação semanal de pós-venda por WhatsApp e benefícios para segunda compra.",
                    "priority": "Alta",
                }
            )

        if items_per_order is not None and items_per_order < 1.4:
            alerts.append(
                {
                    "title": "Baixo mix por pedido",
                    "detail": f"Itens por pedido em {_format_number(items_per_order, 2)}; há espaço para upsell.",
                    "severity": "Média",
                }
            )
            actions.append(
                {
                    "title": "Aumentar itens por atendimento",
                    "detail": "Aplicar combo/kit na frente de caixa e script de venda consultiva.",
                    "priority": "Média",
                }
            )

        if not actions:
            actions.append(
                {
                    "title": "Padronizar rotina comercial da loja",
                    "detail": "Acompanhar ticket, mix e recorrência por vendedor em reunião semanal.",
                    "priority": "Média",
                }
            )

    if not alerts:
        alerts.append(
            {
                "title": "Sem alertas críticos",
                "detail": "Indicadores principais dentro da faixa esperada para o período.",
                "severity": "OK",
            }
        )

    if len(actions) < 3:
        actions.append(
            {
                "title": "Garantir atualização diária da base",
                "detail": "Executar sincronização incremental no início do expediente para manter os dados confiáveis.",
                "priority": "Alta",
            }
        )

    return insights[:4], alerts[:4], actions[:4]


def build_order_velocity_payload(
    source: str,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
) -> dict[str, Any]:
    """Bucket orders by time-to-process (order_date → delivery/shipped status change).
    Since we track flag_open_over_2_days but not a completion timestamp,
    we use the presence of flag_open_over_2_days=0 as proxy for 'processed in time'.
    Returns distribution by processing bucket."""
    if source not in STORE_RULES and source != "all":
        raise ValueError("Fonte desconhecida.")
    source_filter_sql, source_params = _source_scope_sql(source)
    # Use a simpler filter: exclude only cancelled orders (not "em aberto"),
    # so that flagged-open orders are not excluded from the late count.
    not_cancelled = "lower(trim(coalesce(situacao, ''))) NOT LIKE '%cancelado%'"

    date_filter = ""
    date_params: tuple[Any, ...] = ()
    if start_date and end_date:
        date_filter = "AND date(order_date) BETWEEN ? AND ?"
        date_params = (start_date, end_date)

    conn = _db_connect()
    try:
        total_row = conn.execute(
            f"SELECT COUNT(*) AS cnt FROM orders WHERE {source_filter_sql} AND {not_cancelled} {date_filter}",
            source_params + date_params,
        ).fetchone()
        total = int(total_row["cnt"]) if total_row else 0

        open_row = conn.execute(
            f"""SELECT COUNT(*) AS cnt FROM orders
                WHERE {source_filter_sql} AND {not_cancelled} {date_filter}
                  AND flag_open_over_2_days = 1""",
            source_params + date_params,
        ).fetchone()
        open_late = int(open_row["cnt"]) if open_row else 0

        processed_ok = total - open_late
    finally:
        conn.close()

    pct_ok = round(processed_ok / total * 100, 1) if total > 0 else 0.0
    pct_late = round(open_late / total * 100, 1) if total > 0 else 0.0

    return {
        "source": source,
        "periodStart": start_date,
        "periodEnd": end_date,
        "totalOrders": total,
        "processedOk": processed_ok,
        "processedLate": open_late,
        "pctOk": pct_ok,
        "pctLate": pct_late,
        "buckets": [
            {"label": "Processados a tempo", "count": processed_ok, "pct": pct_ok},
            {"label": "Em atraso", "count": open_late, "pct": pct_late},
        ],
    }


def refresh_monthly_aggregates(source: Optional[str] = None) -> int:
    """Recompute monthly_aggregates for all stores (or one store). Returns rows upserted."""
    analytics_filter = _orders_analytics_filter_sql()
    stores = [source] if source and source in STORE_RULES else list(STORE_RULES.keys())
    total_upserted = 0
    computed_at = datetime.now(timezone.utc).replace(tzinfo=None).isoformat()

    conn = _db_connect()
    try:
        for store in stores:
            source_filter_sql, source_params = _source_scope_sql(store)
            rows = conn.execute(
                f"""
                SELECT
                    strftime('%Y', order_date) AS year,
                    strftime('%m', order_date) AS month,
                    SUM(total) AS total_revenue,
                    COUNT(*) AS order_count,
                    AVG(total) AS avg_ticket,
                    COUNT(DISTINCT customer_email) AS unique_customers
                FROM orders
                WHERE {source_filter_sql}
                  AND {analytics_filter}
                  AND order_date IS NOT NULL
                GROUP BY year, month
                """,
                source_params,
            ).fetchall()
            for row in rows:
                conn.execute(
                    """
                    INSERT INTO monthly_aggregates
                        (store_code, year, month, total_revenue, order_count, avg_ticket, unique_customers, computed_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT(store_code, year, month) DO UPDATE SET
                        total_revenue    = excluded.total_revenue,
                        order_count      = excluded.order_count,
                        avg_ticket       = excluded.avg_ticket,
                        unique_customers = excluded.unique_customers,
                        computed_at      = excluded.computed_at
                    """,
                    (
                        store,
                        int(row["year"]),
                        int(row["month"]),
                        float(row["total_revenue"] or 0),
                        int(row["order_count"]),
                        float(row["avg_ticket"] or 0),
                        int(row["unique_customers"]),
                        computed_at,
                    ),
                )
                total_upserted += 1
        conn.commit()
    finally:
        conn.close()

    return total_upserted


def build_freight_analysis_payload(
    source: str,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
) -> dict[str, Any]:
    if source not in STORE_RULES and source != "all":
        raise ValueError("Fonte desconhecida.")
    source_filter_sql, source_params = _source_scope_sql(source)
    analytics_filter = _orders_analytics_filter_sql()

    date_filter = ""
    date_params: tuple[Any, ...] = ()
    if start_date and end_date:
        date_filter = "AND date(order_date) BETWEEN ? AND ?"
        date_params = (start_date, end_date)

    conn = _db_connect()
    try:
        rows = conn.execute(
            f"""
            SELECT
                COALESCE(shipping_method_code, 'desconhecido') AS method,
                COUNT(*) AS order_count,
                0.0 AS avg_shipping_cost,  -- TODO: shipping_cost não disponível no schema atual
                COALESCE(AVG(total), 0.0) AS avg_order_revenue,
                0.0 AS total_shipping_cost,  -- TODO: shipping_cost não disponível no schema atual
                COALESCE(SUM(total), 0.0) AS total_revenue
            FROM orders
            WHERE {source_filter_sql}
              AND {analytics_filter}
              {date_filter}
            GROUP BY shipping_method_code
            ORDER BY total_revenue DESC
            """,
            source_params + date_params,
        ).fetchall()
    finally:
        conn.close()

    items = []
    for row in rows:
        avg_revenue = float(row["avg_order_revenue"] or 0)
        avg_shipping = float(row["avg_shipping_cost"] or 0)
        freight_ratio = round(avg_shipping / avg_revenue * 100, 1) if avg_revenue > 0 else 0.0
        items.append({
            "method": row["method"],
            "orderCount": int(row["order_count"]),
            "avgShippingCost": round(avg_shipping, 2),
            "avgOrderRevenue": round(avg_revenue, 2),
            "totalShippingCost": round(float(row["total_shipping_cost"] or 0), 2),
            "totalRevenue": round(float(row["total_revenue"] or 0), 2),
            "freightRatioPct": freight_ratio,
        })

    return {
        "source": source,
        "periodStart": start_date,
        "periodEnd": end_date,
        "items": items,
    }


def build_management_payload(
    source: str,
    period: str = "30d",
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
) -> dict[str, Any]:
    if source not in STORE_RULES:
        raise ValueError("Fonte desconhecida.")
    period_start, period_end, period_key, period_label, is_custom = _resolve_period_dates(
        period,
        MANAGEMENT_PERIODS,
        start_date,
        end_date,
        datetime.now(),
    )

    span_days = max((period_end - period_start).days + 1, 1)
    previous_end = period_start - timedelta(days=1)
    previous_start = previous_end - timedelta(days=span_days - 1)

    conn = _db_connect()
    try:
        current_snapshot = _snapshot_period(conn, source, period_start, period_end)
        previous_snapshot = _snapshot_period(conn, source, previous_start, previous_end)
        operational_snapshot = _operational_snapshot(conn)
    finally:
        conn.close()

    rule = STORE_RULES[source]
    view_mode = "ecommerce" if rule.get("type") == "ecommerce" else "physical"
    view_mode_label = "E-commerce" if view_mode == "ecommerce" else "Loja Física"

    funnel_summary: Optional[dict[str, Any]] = None
    if view_mode == "ecommerce":
        try:
            funnel_payload = build_funnel_payload(
                source,
                period if period in FUNNEL_PERIODS else "30d",
                start_date=start_date,
                end_date=end_date,
            )
            funnel_steps = funnel_payload.get("steps") or []
            funnel_summary = {
                "available": True,
                "apiUsed": bool(funnel_payload.get("apiUsed")),
                "isSample": bool(funnel_payload.get("isSample")),
                "warning": funnel_payload.get("warning"),
                "sessions": int((funnel_steps[0] or {}).get("value") or 0) if len(funnel_steps) > 0 else 0,
                "engagedSessions": int((funnel_steps[1] or {}).get("value") or 0) if len(funnel_steps) > 1 else 0,
                "addToCart": int((funnel_steps[2] or {}).get("value") or 0) if len(funnel_steps) > 2 else 0,
                "purchases": int((funnel_steps[3] or {}).get("value") or 0) if len(funnel_steps) > 3 else 0,
                "conversionRatePct": round(float(funnel_payload.get("conversion") or 0.0) * 100, 2),
                "engagementRatePct": round(float(funnel_payload.get("engagement") or 0.0) * 100, 2),
                "cartRatePct": round(float(funnel_payload.get("cartRate") or 0.0) * 100, 2),
            }
        except Exception as exc:  # noqa: BLE001
            funnel_summary = {
                "available": False,
                "apiUsed": False,
                "isSample": True,
                "warning": str(exc),
                "sessions": 0,
                "engagedSessions": 0,
                "addToCart": 0,
                "purchases": 0,
                "conversionRatePct": None,
                "engagementRatePct": None,
                "cartRatePct": None,
            }

    latest_sync = get_sync_status(source)
    latest_run = latest_sync.get("latestRun") or {}
    latest_state = latest_sync.get("state") or {}

    kpis: list[dict[str, Any]] = [
        {
            "key": "revenue",
            "label": "Receita do período",
            "type": "currency",
            **_metric_payload(
                float(current_snapshot.get("revenue") or 0.0),
                float(previous_snapshot.get("revenue") or 0.0),
            ),
            "note": f"Período anterior: {_format_currency_brl(float(previous_snapshot.get('revenue') or 0.0))}",
        },
        {
            "key": "orders",
            "label": "Pedidos no período",
            "type": "count",
            **_metric_payload(
                float(current_snapshot.get("orderCount") or 0),
                float(previous_snapshot.get("orderCount") or 0),
            ),
            "note": f"Período anterior: {int(previous_snapshot.get('orderCount') or 0)} pedidos",
        },
        {
            "key": "ticket",
            "label": "Ticket médio",
            "type": "currency",
            **_metric_payload(
                float(current_snapshot["avgTicket"]) if current_snapshot.get("avgTicket") is not None else None,
                float(previous_snapshot["avgTicket"]) if previous_snapshot.get("avgTicket") is not None else None,
            ),
            "note": f"Período anterior: {_format_currency_brl(previous_snapshot.get('avgTicket'))}",
        },
        {
            "key": "customers",
            "label": "Clientes únicos",
            "type": "count",
            **_metric_payload(
                float(current_snapshot.get("uniqueCustomers") or 0),
                float(previous_snapshot.get("uniqueCustomers") or 0),
            ),
            "note": f"Período anterior: {int(previous_snapshot.get('uniqueCustomers') or 0)} clientes",
        },
        {
            "key": "operational_missing_link_marker",
            "label": "Sem vínculo e sem marcador",
            "type": "count",
            **_metric_payload(
                float(operational_snapshot.get("missingLinkMarkerCount") or 0),
                None,
            ),
            "note": "Pedidos sem vínculo com e-commerce e sem tag obrigatória",
        },
        {
            "key": "operational_open_over_2_days",
            "label": f"Pedidos em aberto > {OPEN_ORDER_ALERT_DAYS} dia(s)",
            "type": "count",
            **_metric_payload(
                float(operational_snapshot.get("openOverNDaysCount") or 0),
                None,
            ),
            "note": f"Pedidos com status Em aberto fora do prazo de {OPEN_ORDER_ALERT_DAYS} dia(s)",
        },
    ]

    if view_mode == "ecommerce":
        kpis.extend(
            [
                {
                    "key": "conversion",
                    "label": "Conversão final",
                    "type": "percent",
                    **_metric_payload(
                        float(funnel_summary["conversionRatePct"]) if funnel_summary and funnel_summary.get("conversionRatePct") is not None else None,
                        None,
                    ),
                    "note": "Funil de tráfego e conversão (GA4)",
                },
                {
                    "key": "engaged_sessions",
                    "label": "Sessões engajadas",
                    "type": "count",
                    **_metric_payload(
                        float(funnel_summary["engagedSessions"]) if funnel_summary else None,
                        None,
                    ),
                    "note": "Sessões com interação qualificada no período",
                },
            ]
        )
    else:
        kpis.extend(
            [
                {
                    "key": "items_per_order",
                    "label": "Itens por pedido",
                    "type": "number",
                    **_metric_payload(
                        float(current_snapshot["itemsPerOrder"]) if current_snapshot.get("itemsPerOrder") is not None else None,
                        float(previous_snapshot["itemsPerOrder"]) if previous_snapshot.get("itemsPerOrder") is not None else None,
                    ),
                    "note": "Indicador de mix por atendimento",
                },
                {
                    "key": "repeat_rate",
                    "label": "Taxa de recorrência",
                    "type": "percent",
                    **_metric_payload(
                        float(current_snapshot["repeatRatePct"]) if current_snapshot.get("repeatRatePct") is not None else None,
                        float(previous_snapshot["repeatRatePct"]) if previous_snapshot.get("repeatRatePct") is not None else None,
                    ),
                    "note": "Clientes com mais de 1 compra no período",
                },
            ]
        )

    insights, alerts, actions = _build_management_lists(
        view_mode,
        current_snapshot,
        previous_snapshot,
        period_label,
        funnel_summary,
        operational_snapshot,
    )

    return {
        "source": source,
        "storeLabel": rule.get("label"),
        "storeType": rule.get("type"),
        "viewMode": view_mode,
        "viewModeLabel": view_mode_label,
        "period": period_key,
        "periodLabel": period_label,
        "periodIsCustom": is_custom,
        "periodStart": period_start.isoformat(),
        "periodEnd": period_end.isoformat(),
        "previousPeriodStart": previous_start.isoformat(),
        "previousPeriodEnd": previous_end.isoformat(),
        "kpis": kpis,
        "insights": insights,
        "alerts": alerts,
        "actions": actions,
        "topProducts": current_snapshot.get("topProducts") or [],
        "topStates": current_snapshot.get("topStates") or [],
        "paymentMix": current_snapshot.get("paymentMix") or [],
        "shippingMix": current_snapshot.get("shippingMix") or [],
        "operational": operational_snapshot,
        "funnel": funnel_summary,
        "quality": {
            "syncStatus": latest_run.get("status") or "unknown",
            "syncAt": latest_run.get("finished_at") or latest_run.get("started_at") or latest_state.get("last_success_at"),
            "lastSuccessAt": latest_state.get("last_success_at"),
            "funnelStatus": (
                "na"
                if view_mode != "ecommerce"
                else ("sample" if (funnel_summary or {}).get("isSample") else "api")
            ),
        },
    }


def get_sync_status(source: Optional[str] = None) -> dict[str, Any]:
    ensure_database()
    conn = _db_connect()
    try:
        if source and source != "all":
            latest_run = conn.execute(
                "SELECT id, store_code, mode, requested_from, requested_to, started_at, finished_at, status, "
                "orders_found, orders_synced, errors_count, details "
                "FROM sync_runs WHERE store_code = ? ORDER BY id DESC LIMIT 1",
                (source,),
            ).fetchone()
            state = conn.execute(
                "SELECT store_code, last_success_at, last_order_date, last_run_id, updated_at "
                "FROM sync_state WHERE store_code = ?",
                (source,),
            ).fetchone()
            return {
                "source": source,
                "latestRun": dict(latest_run) if latest_run else None,
                "state": dict(state) if state else None,
            }

        latest_runs = conn.execute(
            "SELECT id, store_code, mode, requested_from, requested_to, started_at, finished_at, status, "
            "orders_found, orders_synced, errors_count, details "
            "FROM sync_runs ORDER BY id DESC LIMIT 20"
        ).fetchall()

        states = conn.execute(
            "SELECT store_code, last_success_at, last_order_date, last_run_id, updated_at FROM sync_state"
        ).fetchall()
        latest_run = latest_runs[0] if latest_runs else None
        latest_state = conn.execute(
            "SELECT store_code, last_success_at, last_order_date, last_run_id, updated_at "
            "FROM sync_state "
            "ORDER BY COALESCE(last_success_at, updated_at) DESC, updated_at DESC "
            "LIMIT 1"
        ).fetchone()

        return {
            "source": source or "all",
            "latestRun": dict(latest_run) if latest_run else None,
            "state": dict(latest_state) if latest_state else None,
            "latestRuns": [dict(row) for row in latest_runs],
            "states": [dict(row) for row in states],
        }
    finally:
        conn.close()


def build_bigb_clients_payload() -> dict:
    """
    Returns Big B clients with days since last purchase.
    Always calculated relative to today (date.today()).
    Source is always 'bigb' — no parameter needed.
    """
    today = date.today()

    scope_sql, scope_params = _source_scope_sql("bigb", alias="o")
    scope_sql_inner, scope_params_inner = _source_scope_sql("bigb", alias="o2")

    sql = f"""
        SELECT
            (
                SELECT o2.customer_name
                FROM orders o2
                WHERE o2.customer_email = o.customer_email
                  AND {scope_sql_inner}
                  AND lower(trim(coalesce(o2.situacao, ''))) NOT LIKE '%cancelado%'
                ORDER BY date(o2.order_date) DESC
                LIMIT 1
            ) AS name,
            o.customer_email                        AS email,
            MAX(date(o.order_date))                 AS last_purchase_date,
            SUM(o.total)                            AS total_revenue
        FROM orders o
        WHERE {scope_sql}
          AND lower(trim(coalesce(o.situacao, ''))) NOT LIKE '%cancelado%'
        GROUP BY o.customer_email
        ORDER BY last_purchase_date ASC
    """

    with _db_connect() as conn:
        # params order must match SQL: inner subquery first, outer WHERE second
        rows = conn.execute(sql, scope_params_inner + scope_params).fetchall()

    clients = []
    alert_count = 0
    for row in rows:
        if not row["last_purchase_date"]:
            continue
        # Skip rows with null/empty email
        email = row["email"] or ""
        if not email.strip():
            continue
        last_date = date.fromisoformat(row["last_purchase_date"])
        days_since = (today - last_date).days
        if days_since >= 40:
            alert_count += 1
        clients.append({
            "name": row["name"] or "",
            "email": email,
            "lastPurchaseDate": row["last_purchase_date"],
            "daysSince": days_since,
            "totalRevenue": float(row["total_revenue"] or 0),
        })

    # Sort by days descending (most overdue first)
    clients.sort(key=lambda c: c["daysSince"], reverse=True)

    return {
        "clients": clients,
        "alertCount": alert_count,
    }


# ── Monthly KPIs (new for Dash_Final) ──────────────────────────────

def build_monthly_kpis(source: str, year: int, month: int) -> dict:
    """
    Returns KPI card data for a specific month and the delta vs previous month.
    Used by the 3-card row in the new dashboard.
    """
    db_path = DB_PATH
    excluded = list(EXCLUDED_ORDER_SITUACAO_TOKENS)
    excl_placeholders = ",".join(f"'{s}'" for s in excluded)

    def _query_month(yr: int, mo: int) -> dict:
        month_str = f"{yr}-{mo:02d}"
        with sqlite3.connect(db_path) as conn:
            conn.row_factory = sqlite3.Row
            # Get store rules for this source
            store_rules = STORE_RULES.get(source)
            if store_rules is None:
                return {"revenue": None, "orders": None}

            if store_rules["type"] == "ecommerce":
                id_filter = "AND s.ecommerce_name = :ecommerce_name"
                params: dict = {"month": month_str, "ecommerce_name": store_rules["ecommerce_name"]}
            else:
                id_filter = "AND s.marker_tag = :marker_tag"
                params = {"month": month_str, "marker_tag": store_rules["marker_tag"]}

            if excl_placeholders:
                excl_clause = f"AND lower(trim(coalesce(o.situacao, ''))) NOT IN ({excl_placeholders})"
            else:
                excl_clause = ""

            row = conn.execute(f"""
                SELECT
                    COALESCE(SUM(o.total), 0) AS revenue,
                    COUNT(*) AS orders
                FROM orders o
                JOIN stores s ON o.store_code = s.code
                WHERE strftime('%Y-%m', o.order_date) = :month
                {id_filter}
                {excl_clause}
            """, params).fetchone()
            return {"revenue": float(row["revenue"]) if row else 0.0,
                    "orders": int(row["orders"]) if row else 0}

    def _query_target(yr: int, mo: int) -> float | None:
        with sqlite3.connect(db_path) as conn:
            row = conn.execute(
                "SELECT target_revenue FROM revenue_targets_monthly WHERE store_code = ? AND year = ? AND month = ?",
                (source, yr, mo),
            ).fetchone()
            return float(row[0]) if row and row[0] is not None else None

    curr = _query_month(year, month)
    prev_year, prev_month = year - 1, month   # same month, previous year (seasonal market)
    prev = _query_month(prev_year, prev_month)
    target = _query_target(year, month)

    rev = curr["revenue"]
    prev_rev = prev["revenue"]
    orders = curr["orders"]
    prev_orders = prev["orders"]

    ticket = (rev / orders) if orders else None
    prev_ticket = (prev_rev / prev_orders) if prev_orders else None

    def pct_delta(curr_v, prev_v):
        if prev_v and prev_v != 0:
            return round((curr_v - prev_v) / prev_v * 100, 1)
        return None

    return {
        "source": source,
        "year": year,
        "month": month,
        "revenue": rev,
        "orders": orders,
        "ticket": round(ticket, 2) if ticket else None,
        "revenue_delta_pct": pct_delta(rev, prev_rev),
        "orders_delta_abs": orders - prev_orders if prev_orders is not None else None,
        "ticket_delta_pct": pct_delta(ticket, prev_ticket) if ticket and prev_ticket else None,
        "target": target,
        "gap_to_target": round(target - rev, 2) if target is not None else None,
    }
