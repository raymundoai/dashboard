from __future__ import annotations

import sqlite3
from pathlib import Path

from openpyxl import Workbook

import tiny_bi


def _write_brew_history_workbook(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Planilha1"

    worksheet["B1"] = "FATURAMENTO E-COMMERCE BREW"
    for column, year in zip(("C", "D", "E", "F", "G", "H"), (2021, 2022, 2023, 2024, 2025, 2026), strict=True):
        worksheet[f"{column}3"] = year

    rows = [
        ("JANEIRO", 101.0, 201.0, 301.0, 401.0, 501.0, 601.0),
        ("FEVEREIRO", 102.0, 202.0, 302.0, 402.0, 502.0, 602.0),
        ("MARÇO", 103.0, 203.0, 303.0, 403.0, 503.0, 603.0),
        ("ABRIL", 104.0, 204.0, 304.0, 404.0, 504.0, 604.0),
        ("MAIO", 105.0, 205.0, 305.0, 405.0, 505.0, 605.0),
        ("JUNHO", 106.0, 206.0, 306.0, 406.0, 506.0, 606.0),
        ("JULHO", 107.0, 207.0, 307.0, 407.0, 507.0, 607.0),
        ("AGOSTO", 108.0, 208.0, 308.0, 408.0, 508.0, 608.0),
        ("SETEMBRO", 109.0, 209.0, 309.0, 409.0, 509.0, 609.0),
        ("OUTUBRO", 110.0, 210.0, 310.0, 410.0, 510.0, 610.0),
        ("NOVEMBRO", 111.0, 211.0, 311.0, 411.0, 511.0, 611.0),
        ("DEZEMBRO", 112.0, 212.0, 312.0, 412.0, 512.0, 612.0),
    ]

    for row_index, row in enumerate(rows, start=4):
        worksheet.cell(row=row_index, column=2, value=row[0])
        for column_index, value in enumerate(row[1:], start=3):
            worksheet.cell(row=row_index, column=column_index, value=value)

    workbook.save(path)


def test_import_brew_history_overwrites_only_2021_2024_and_preserves_targets(client, tmp_path):
    workbook_path = tmp_path / "brew-history.xlsx"
    _write_brew_history_workbook(workbook_path)

    conn = sqlite3.connect(tiny_bi.DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        conn.execute("DELETE FROM revenue_targets_monthly WHERE store_code = 'brew' AND year BETWEEN 2021 AND 2026")
        conn.execute(
            """
            INSERT INTO revenue_targets_monthly
                (store_code, year, month, target_revenue, realized_revenue, source_file)
            VALUES ('brew', 2021, 1, NULL, 999.0, 'legacy.xlsx')
            """
        )
        conn.execute(
            """
            INSERT INTO revenue_targets_monthly
                (store_code, year, month, target_revenue, realized_revenue, source_file)
            VALUES ('brew', 2024, 2, 555.0, 888.0, 'legacy.xlsx')
            """
        )
        conn.execute(
            """
            INSERT INTO revenue_targets_monthly
                (store_code, year, month, target_revenue, realized_revenue, source_file)
            VALUES ('brew', 2025, 1, 999.0, 777.0, 'future.xlsx')
            """
        )
        conn.execute(
            """
            INSERT INTO revenue_targets_monthly
                (store_code, year, month, target_revenue, realized_revenue, source_file)
            VALUES ('brew', 2026, 1, 111.0, NULL, 'future.xlsx')
            """
        )
        conn.commit()
    finally:
        conn.close()

    result = tiny_bi.import_brew_monthly_history_workbook(workbook_path, start_year=2021, end_year=2024)

    assert result["status"] == "success"
    assert result["rowsUpserted"] == 48
    assert result["rowsInserted"] == 46
    assert result["rowsOverwritten"] == 2
    assert result["targetsPreserved"] == 1
    assert result["yearsImported"] == [2021, 2022, 2023, 2024]

    conn = sqlite3.connect(tiny_bi.DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        jan_2021 = conn.execute(
            """
            SELECT target_revenue, realized_revenue, source_file
            FROM revenue_targets_monthly
            WHERE store_code = 'brew' AND year = 2021 AND month = 1
            """
        ).fetchone()
        assert jan_2021["target_revenue"] is None
        assert jan_2021["realized_revenue"] == 101.0
        assert jan_2021["source_file"] == workbook_path.name

        feb_2024 = conn.execute(
            """
            SELECT target_revenue, realized_revenue
            FROM revenue_targets_monthly
            WHERE store_code = 'brew' AND year = 2024 AND month = 2
            """
        ).fetchone()
        assert feb_2024["target_revenue"] == 555.0
        assert feb_2024["realized_revenue"] == 402.0

        jul_2023 = conn.execute(
            """
            SELECT realized_revenue
            FROM revenue_targets_monthly
            WHERE store_code = 'brew' AND year = 2023 AND month = 7
            """
        ).fetchone()
        assert jul_2023["realized_revenue"] == 307.0

        jan_2025 = conn.execute(
            """
            SELECT target_revenue, realized_revenue, source_file
            FROM revenue_targets_monthly
            WHERE store_code = 'brew' AND year = 2025 AND month = 1
            """
        ).fetchone()
        assert jan_2025["target_revenue"] == 999.0
        assert jan_2025["realized_revenue"] == 777.0
        assert jan_2025["source_file"] == "future.xlsx"

        jan_2026 = conn.execute(
            """
            SELECT target_revenue, realized_revenue, source_file
            FROM revenue_targets_monthly
            WHERE store_code = 'brew' AND year = 2026 AND month = 1
            """
        ).fetchone()
        assert jan_2026["target_revenue"] == 111.0
        assert jan_2026["realized_revenue"] is None
        assert jan_2026["source_file"] == "future.xlsx"
    finally:
        conn.execute("DELETE FROM revenue_targets_monthly WHERE store_code = 'brew' AND year BETWEEN 2021 AND 2026")
        conn.commit()
        conn.close()
